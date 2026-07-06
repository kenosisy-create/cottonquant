"""R53 CF fundamental manual-input quality and observation."""

from __future__ import annotations

import csv
import json
import uuid
import warnings as py_warnings
from dataclasses import dataclass
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from cotton_factor.common.exceptions import ResearchWorkbenchError
from cotton_factor.common.paths import data_dir, reports_dir
from cotton_factor.common.time import utc_now

PRODUCT_CODE = "CF"
FUNDAMENTAL_OBSERVATION_VERSION = "R53_fundamental_observation_v1"
OUTPUT_DIR = "fundamentals"
REPORT_DIR = "fundamentals"
INPUT_SUFFIXES = {".csv", ".xlsx", ".xls"}
TEXTILE_METRIC_HEADERS = ("日均", "周均", "月均")
TEXTILE_FILE_KEYWORDS = (
    "TTEB",
    "开工",
    "负荷",
    "纱厂",
    "织厂",
    "棉纱",
    "坯布",
    "纺企",
)
TEXTILE_CANONICAL_INDICATOR_MAP = {
    "纱线综合库存": (
        "纺企棉纱库存",
        "2025-12-19 起 TTEB 原始表口径名为纱线综合库存，按纺企棉纱库存延续口径处理",
    ),
}
IFIND_METADATA_KEYS = {
    "频率": "frequency",
    "单位": "unit",
    "指标ID": "indicator_id",
    "来源": "source_name",
    "更新时间": "update_time",
}
HUMAN_REVIEW_REQUIRED = (
    "official_fundamental_field_interpretation",
    "basis_spot_source_and_region",
    "inventory_source_and_unit",
    "warehouse_receipt_quantity_source",
    "import_period_and_unit",
    "textile_chain_missing_data",
    "fundamental_signal_rule_before_use",
)


@dataclass(frozen=True)
class FundamentalObservationWarningRecord:
    """Warning row for R53 fundamental observation."""

    severity: str
    warning_code: str
    message: str
    affected_count: int = 0
    human_review_required: tuple[str, ...] = ()

    def to_summary(self) -> dict[str, object]:
        """Return a JSON-serializable warning row."""
        return {
            "severity": self.severity,
            "warning_code": self.warning_code,
            "message": self.message,
            "affected_count": self.affected_count,
            "human_review_required": list(self.human_review_required),
        }


@dataclass(frozen=True)
class FundamentalObservationDatasetSummary:
    """Dataset-level R53 quality summary."""

    dataset_type: str
    status: str
    row_count: int
    date_start: str | None
    date_end: str | None
    source_files: tuple[str, ...]
    notes: str

    def to_summary(self) -> dict[str, object]:
        """Return a JSON-serializable dataset summary."""
        return {
            "dataset_type": self.dataset_type,
            "status": self.status,
            "row_count": self.row_count,
            "date_start": self.date_start,
            "date_end": self.date_end,
            "source_files": list(self.source_files),
            "notes": self.notes,
        }


@dataclass(frozen=True)
class ResearchFundamentalObservationResult:
    """Result of building R53 fundamental observation artifacts."""

    product_code: str
    run_id: str
    status: str
    data_asof: date | None
    source_dir: Path
    input_file_count: int
    dataset_summaries: tuple[FundamentalObservationDatasetSummary, ...]
    warning_records: tuple[FundamentalObservationWarningRecord, ...]
    inventory_path: Path
    basis_path: Path
    spot_path: Path
    warehouse_receipt_path: Path
    import_path: Path
    textile_chain_path: Path
    field_metadata_csv_path: Path
    quality_csv_path: Path
    warning_csv_path: Path
    json_path: Path
    manifest_path: Path
    markdown_path: Path
    human_review_required: tuple[str, ...]

    @property
    def passed(self) -> bool:
        """R53 passes when usable observation artifacts are written."""
        return self.status in {"OBSERVATION_READY_WITH_WARNINGS", "NO_USABLE_INPUT"}

    @property
    def warning_count(self) -> int:
        """Return non-info warning count."""
        return sum(1 for warning in self.warning_records if warning.severity != "INFO")

    def to_summary(self) -> dict[str, object]:
        """Return a compact CLI summary."""
        return {
            "product_code": self.product_code,
            "run_id": self.run_id,
            "status": self.status,
            "passed": self.passed,
            "data_asof": None if self.data_asof is None else self.data_asof.isoformat(),
            "source_dir": str(self.source_dir),
            "input_file_count": self.input_file_count,
            "dataset_summaries": [
                summary.to_summary() for summary in self.dataset_summaries
            ],
            "warning_count": self.warning_count,
            "warnings": [warning.to_summary() for warning in self.warning_records],
            "inventory_path": str(self.inventory_path),
            "basis_path": str(self.basis_path),
            "spot_path": str(self.spot_path),
            "warehouse_receipt_path": str(self.warehouse_receipt_path),
            "import_path": str(self.import_path),
            "textile_chain_path": str(self.textile_chain_path),
            "field_metadata_csv_path": str(self.field_metadata_csv_path),
            "quality_csv_path": str(self.quality_csv_path),
            "warning_csv_path": str(self.warning_csv_path),
            "json_path": str(self.json_path),
            "manifest_path": str(self.manifest_path),
            "markdown_path": str(self.markdown_path),
            "fundamental_signal_status": "not_connected",
            "human_review_required": list(self.human_review_required),
        }


def build_cf_fundamental_observation(
    *,
    source_dir: Path | None = None,
    output_dir: Path | None = None,
    report_output_dir: Path | None = None,
    run_id: str | None = None,
) -> ResearchFundamentalObservationResult:
    """Build R53 quality tables and Chinese observation report from manual inputs."""
    observation_run_id = run_id or _default_run_id()
    input_dir = source_dir or data_dir() / "incoming" / PRODUCT_CODE / "fundamentals" / "manual"
    if input_dir.exists() and not input_dir.is_dir():
        raise ResearchWorkbenchError(f"fundamental source path is not a directory: {input_dir}")
    input_dir.mkdir(parents=True, exist_ok=True)
    input_files = _input_files(input_dir)

    warnings: list[FundamentalObservationWarningRecord] = []
    long_frames: list[pd.DataFrame] = []
    textile_frames: list[pd.DataFrame] = []
    for path in input_files:
        if path.suffix.lower() == ".csv":
            try:
                long_frames.append(_read_ifind_wide_csv(path))
            except ResearchWorkbenchError as exc:
                warnings.append(
                    FundamentalObservationWarningRecord(
                        severity="WARN",
                        warning_code="FUNDAMENTAL_FILE_PARSE_FAILED",
                        message=str(exc),
                        affected_count=1,
                        human_review_required=("official_fundamental_field_interpretation",),
                    )
                )
        elif path.suffix.lower() in {".xlsx", ".xls"}:
            textile_frame, textile_warnings = _read_textile_chain_xlsx(path)
            if not textile_frame.empty:
                textile_frames.append(textile_frame)
                warnings.extend(textile_warnings)
                continue
            warnings.extend(textile_warnings)
            frame, xlsx_warnings = _read_ifind_narrow_xlsx(path)
            if not frame.empty:
                long_frames.append(frame)
            warnings.extend(xlsx_warnings)

    long = pd.concat(long_frames, ignore_index=True) if long_frames else _empty_long_frame()
    textile_chain = (
        pd.concat(textile_frames, ignore_index=True)
        if textile_frames
        else _empty_textile_chain_frame()
    )
    inventory = _inventory_rows(long)
    spot = _spot_rows(long)
    basis = _basis_rows(long)
    warehouse_receipt = _warehouse_receipt_rows(long)
    import_rows = _import_rows(long)
    field_metadata = _field_metadata_rows(long, textile_chain=textile_chain)
    warnings.extend(
        _dataset_warnings(
            input_files=input_files,
            long=long,
            basis=basis,
            textile_chain=textile_chain,
        )
    )
    dataset_summaries = _dataset_summaries(
        inventory=inventory,
        basis=basis,
        spot=spot,
        warehouse_receipt=warehouse_receipt,
        import_rows=import_rows,
        textile_chain=textile_chain,
        input_files=input_files,
    )
    data_asof = _max_date(
        (inventory, basis, spot, warehouse_receipt, import_rows, textile_chain)
    )
    status = "OBSERVATION_READY_WITH_WARNINGS" if data_asof is not None else "NO_USABLE_INPUT"
    result = ResearchFundamentalObservationResult(
        product_code=PRODUCT_CODE,
        run_id=observation_run_id,
        status=status,
        data_asof=data_asof,
        source_dir=input_dir,
        input_file_count=len(input_files),
        dataset_summaries=tuple(dataset_summaries),
        warning_records=tuple(warnings),
        inventory_path=_inventory_path(output_dir),
        basis_path=_basis_path(output_dir),
        spot_path=_spot_path(output_dir),
        warehouse_receipt_path=_warehouse_receipt_path(output_dir),
        import_path=_import_path(output_dir),
        textile_chain_path=_textile_chain_path(output_dir),
        field_metadata_csv_path=_field_metadata_csv_path(output_dir),
        quality_csv_path=_quality_csv_path(output_dir),
        warning_csv_path=_warning_csv_path(report_output_dir),
        json_path=_json_path(output_dir),
        manifest_path=_manifest_path(output_dir),
        markdown_path=_markdown_path(report_output_dir),
        human_review_required=HUMAN_REVIEW_REQUIRED,
    )
    _write_outputs(
        result=result,
        inventory=inventory,
        basis=basis,
        spot=spot,
        warehouse_receipt=warehouse_receipt,
        import_rows=import_rows,
        textile_chain=textile_chain,
        field_metadata=field_metadata,
    )
    return result


def _input_files(source_dir: Path) -> tuple[Path, ...]:
    return tuple(
        sorted(
            path
            for path in source_dir.iterdir()
            if path.is_file() and path.suffix.lower() in INPUT_SUFFIXES
            and not path.name.startswith("~$")
        )
    )


def _read_ifind_wide_csv(path: Path) -> pd.DataFrame:
    frame: pd.DataFrame | None = None
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
        try:
            frame = pd.read_csv(path, encoding=encoding, header=None)
            break
        except Exception as exc:  # pragma: no cover - exercised only on bad files.
            last_error = exc
    if frame is None:
        raise ResearchWorkbenchError(f"cannot read iFinD CSV {path}: {last_error}")
    if len(frame) < 7 or frame.shape[1] < 2:
        raise ResearchWorkbenchError(f"iFinD CSV has no usable data rows: {path}")

    # iFinD 手工导出的宽表前 6 行是元数据，后续才是日期序列。
    indicator_names = [str(value) for value in frame.iloc[0].tolist()]
    metadata = _metadata_by_column(frame)
    data = frame.iloc[6:].copy()
    date_values = pd.to_datetime(data.iloc[:, 0], errors="coerce").dt.date
    rows: list[dict[str, object]] = []
    for col_index in range(1, frame.shape[1]):
        indicator_name = indicator_names[col_index]
        meta = metadata.get(col_index, {})
        values = pd.to_numeric(data.iloc[:, col_index], errors="coerce")
        for row_index, value in enumerate(values):
            trade_date = date_values.iloc[row_index]
            if pd.isna(value) or pd.isna(trade_date):
                continue
            rows.append(
                {
                    "trade_date": trade_date,
                    "indicator_name": indicator_name,
                    "indicator_value": float(value),
                    "unit": meta.get("unit"),
                    "frequency": meta.get("frequency"),
                    "indicator_id": meta.get("indicator_id"),
                    "source_name": meta.get("source_name") or "iFinD",
                    "update_time": meta.get("update_time"),
                    "source_file": str(path),
                }
            )
    if not rows:
        raise ResearchWorkbenchError(f"iFinD CSV has only metadata or empty values: {path}")
    return pd.DataFrame(rows)


def _metadata_by_column(frame: pd.DataFrame) -> dict[int, dict[str, object]]:
    metadata: dict[int, dict[str, object]] = {}
    for row_index in range(1, min(6, len(frame))):
        raw_key = str(frame.iloc[row_index, 0])
        normalized_key = IFIND_METADATA_KEYS.get(raw_key)
        if normalized_key is None:
            continue
        for col_index in range(1, frame.shape[1]):
            value = frame.iloc[row_index, col_index]
            if pd.isna(value):
                continue
            metadata.setdefault(col_index, {})[normalized_key] = value
    return metadata


def _read_textile_chain_xlsx(
    path: Path,
) -> tuple[pd.DataFrame, list[FundamentalObservationWarningRecord]]:
    if not _looks_like_textile_chain_file(path):
        return _empty_textile_chain_frame(), []
    try:
        with py_warnings.catch_warnings():
            py_warnings.filterwarnings(
                "ignore",
                message="Workbook contains no default style.*",
                category=UserWarning,
            )
            workbook = load_workbook(path, data_only=False, read_only=True)
    except Exception as exc:  # pragma: no cover - depends on corrupted workbook.
        return (
            _empty_textile_chain_frame(),
            [
                FundamentalObservationWarningRecord(
                    severity="WARN",
                    warning_code="TEXTILE_CHAIN_XLSX_PARSE_FAILED",
                    message=f"纺织链 Excel 文件无法读取：{path}；{exc}",
                    affected_count=1,
                    human_review_required=("official_fundamental_field_interpretation",),
                )
            ],
        )

    frames: list[pd.DataFrame] = []
    for sheet in workbook.worksheets:
        rows = [tuple(row) for row in sheet.iter_rows(values_only=True)]
        frame = _parse_textile_block_sheet(rows=rows, path=path, sheet_title=sheet.title)
        if not frame.empty:
            frames.append(frame)
    workbook.close()
    if frames:
        return pd.concat(frames, ignore_index=True), []
    return (
        _empty_textile_chain_frame(),
        [
            FundamentalObservationWarningRecord(
                severity="WARN",
                warning_code="TEXTILE_CHAIN_XLSX_UNSUPPORTED_LAYOUT",
                message=f"纺织链 Excel 文件存在内容但未识别到 TTEB 分块结构：{path}",
                affected_count=1,
                human_review_required=("official_fundamental_field_interpretation",),
            )
        ],
    )


def _looks_like_textile_chain_file(path: Path) -> bool:
    text = path.name
    return any(keyword in text for keyword in TEXTILE_FILE_KEYWORDS)


def _parse_textile_block_sheet(
    *,
    rows: list[tuple[object, ...]],
    path: Path,
    sheet_title: str,
) -> pd.DataFrame:
    if not rows:
        return _empty_textile_chain_frame()
    groups = _textile_block_groups(rows[0])
    if not groups:
        return _empty_textile_chain_frame()

    parsed_rows: list[dict[str, object]] = []
    for product_col, product_label, metric_columns in groups:
        date_col = product_col + 1
        for row in rows[1:]:
            if date_col >= len(row):
                continue
            trade_date = pd.to_datetime(row[date_col], errors="coerce")
            if pd.isna(trade_date):
                continue
            row_indicator = _cell_text(row[product_col]) if product_col < len(row) else None
            raw_indicator_name = row_indicator or product_label
            if raw_indicator_name == "产品":
                continue
            indicator_name, continuation_remark = _canonical_textile_indicator_name(
                raw_indicator_name
            )
            for metric_col, metric_name in metric_columns:
                if metric_col >= len(row):
                    continue
                value = pd.to_numeric(row[metric_col], errors="coerce")
                if pd.isna(value):
                    continue
                unit = _textile_unit(
                    indicator_name=indicator_name,
                    sheet_title=sheet_title,
                    path=path,
                )
                parsed_rows.append(
                    {
                        "trade_date": trade_date.date(),
                        "product_code": PRODUCT_CODE,
                        "indicator_name": indicator_name,
                        "raw_indicator_name": raw_indicator_name,
                        "metric_name": metric_name,
                        "indicator_value": float(value),
                        "unit": unit,
                        "frequency": "周",
                        "indicator_id": (
                            f"TTEB:{sheet_title}:{raw_indicator_name}:{metric_name}"
                        ),
                        "source_name": "TTEB",
                        "source_file": str(path),
                        "source_sheet": sheet_title,
                        "data_quality_flag": "REVIEW_REQUIRED",
                        "human_review_required": True,
                        # TTEB 原始表为人工下载的块状表，先作为观察项保留，不进入信号。
                        "remark": _textile_remark(continuation_remark),
                    }
                )
    if not parsed_rows:
        return _empty_textile_chain_frame()
    return (
        pd.DataFrame(parsed_rows)
        .sort_values(["trade_date", "indicator_name", "metric_name"])
        .reset_index(drop=True)
    )


def _canonical_textile_indicator_name(raw_indicator_name: str) -> tuple[str, str | None]:
    """把 TTEB 原始指标名映射到研究口径，同时保留原始名用于审计。"""
    mapped = TEXTILE_CANONICAL_INDICATOR_MAP.get(raw_indicator_name)
    if mapped is None:
        return raw_indicator_name, None
    return mapped


def _textile_remark(continuation_remark: str | None) -> str:
    base = "TTEB block workbook; textile-chain unit/source require human review"
    if continuation_remark is None:
        return base
    return f"{base}; {continuation_remark}"


def _textile_block_groups(
    header: tuple[object, ...],
) -> list[tuple[int, str, tuple[tuple[int, str], ...]]]:
    groups: list[tuple[int, str, tuple[tuple[int, str], ...]]] = []
    for col_index in range(max(0, len(header) - 2)):
        product_label = _cell_text(header[col_index])
        date_label = _cell_text(header[col_index + 1])
        if not product_label or date_label != "日期":
            continue
        metric_columns: list[tuple[int, str]] = []
        for metric_offset in range(2, 5):
            metric_col = col_index + metric_offset
            if metric_col >= len(header):
                continue
            metric_name = _cell_text(header[metric_col])
            if metric_name in TEXTILE_METRIC_HEADERS:
                metric_columns.append((metric_col, metric_name))
        if metric_columns:
            groups.append((col_index, product_label, tuple(metric_columns)))
    return groups


def _cell_text(value: object) -> str | None:
    if value is None or pd.isna(value):
        return None
    text = str(value).strip()
    return text or None


def _textile_unit(*, indicator_name: str, sheet_title: str, path: Path) -> str:
    text = f"{path.name} {sheet_title} {indicator_name}"
    if "%" in text or "负荷" in text or "开工" in text:
        return "%"
    if "库存" in text or "天数" in text:
        return "天"
    return "HUMAN_REVIEW_REQUIRED"


def _xlsx_warnings(path: Path) -> list[FundamentalObservationWarningRecord]:
    warnings: list[FundamentalObservationWarningRecord] = []
    try:
        with py_warnings.catch_warnings():
            py_warnings.filterwarnings(
                "ignore",
                message="Workbook contains no default style.*",
                category=UserWarning,
            )
            workbook = load_workbook(path, data_only=False, read_only=True)
    except Exception as exc:  # pragma: no cover - depends on corrupted workbook.
        return [
            FundamentalObservationWarningRecord(
                severity="WARN",
                warning_code="FUNDAMENTAL_XLSX_PARSE_FAILED",
                message=f"Excel 文件无法读取：{path}；{exc}",
                affected_count=1,
                human_review_required=("official_fundamental_field_interpretation",),
            )
        ]
    for sheet in workbook.worksheets:
        non_empty = [
            cell
            for row in sheet.iter_rows(values_only=True)
            for cell in row
            if cell is not None and str(cell).strip()
        ]
        if not non_empty:
            warnings.append(
                FundamentalObservationWarningRecord(
                    severity="WARN",
                    warning_code="FUNDAMENTAL_XLSX_EMPTY",
                    message=f"Excel 工作表为空：{path} / {sheet.title}",
                    affected_count=1,
                    human_review_required=("official_fundamental_field_interpretation",),
                )
            )
        elif len(non_empty) == 1 and str(non_empty[0]).startswith("="):
            warnings.append(
                FundamentalObservationWarningRecord(
                    severity="WARN",
                    warning_code="IMPORT_INPUT_NOT_REFRESHED",
                    message=(
                        f"进口 Excel 仅发现未刷新公式 {non_empty[0]!r}：{path}。"
                        "请在 iFinD Excel 插件中刷新并另存为数值。"
                    ),
                    affected_count=1,
                    human_review_required=("import_period_and_unit",),
                )
            )
        else:
            warnings.append(
                FundamentalObservationWarningRecord(
                    severity="WARN",
                    warning_code="FUNDAMENTAL_XLSX_UNSUPPORTED_LAYOUT",
                    message=f"Excel 文件存在内容但尚未支持自动规范化：{path} / {sheet.title}",
                    affected_count=1,
                    human_review_required=("official_fundamental_field_interpretation",),
                )
            )
    workbook.close()
    return warnings


def _read_ifind_narrow_xlsx(
    path: Path,
) -> tuple[pd.DataFrame, list[FundamentalObservationWarningRecord]]:
    try:
        with py_warnings.catch_warnings():
            py_warnings.filterwarnings(
                "ignore",
                message="Workbook contains no default style.*",
                category=UserWarning,
            )
            workbook = load_workbook(path, data_only=False, read_only=True)
    except Exception as exc:  # pragma: no cover - depends on corrupted workbook.
        return (
            _empty_long_frame(),
            [
                FundamentalObservationWarningRecord(
                    severity="WARN",
                    warning_code="FUNDAMENTAL_XLSX_PARSE_FAILED",
                    message=f"Excel 文件无法读取：{path}；{exc}",
                    affected_count=1,
                    human_review_required=("official_fundamental_field_interpretation",),
                )
            ],
        )

    all_warnings: list[FundamentalObservationWarningRecord] = []
    frames: list[pd.DataFrame] = []
    for sheet in workbook.worksheets:
        rows = [tuple(row) for row in sheet.iter_rows(values_only=True)]
        non_empty = [
            cell
            for row in rows
            for cell in row
            if cell is not None and str(cell).strip()
        ]
        if not non_empty:
            all_warnings.append(
                FundamentalObservationWarningRecord(
                    severity="WARN",
                    warning_code="FUNDAMENTAL_XLSX_EMPTY",
                    message=f"Excel 工作表为空：{path} / {sheet.title}",
                    affected_count=1,
                    human_review_required=("official_fundamental_field_interpretation",),
                )
            )
            continue
        frame = _parse_ifind_narrow_sheet(rows=rows, path=path)
        if not frame.empty:
            frames.append(frame)
            continue
        if len(non_empty) == 1 and str(non_empty[0]).startswith("="):
            all_warnings.append(
                FundamentalObservationWarningRecord(
                    severity="WARN",
                    warning_code="IMPORT_INPUT_NOT_REFRESHED",
                    message=(
                        f"进口 Excel 仅发现未刷新公式 {non_empty[0]!r}：{path}。"
                        "请在 iFinD Excel 插件中刷新并另存为数值。"
                    ),
                    affected_count=1,
                    human_review_required=("import_period_and_unit",),
                )
            )
        else:
            all_warnings.append(
                FundamentalObservationWarningRecord(
                    severity="WARN",
                    warning_code="FUNDAMENTAL_XLSX_UNSUPPORTED_LAYOUT",
                    message=f"Excel 文件存在内容但尚未支持自动规范化：{path} / {sheet.title}",
                    affected_count=1,
                    human_review_required=("official_fundamental_field_interpretation",),
                )
            )
    workbook.close()
    long = pd.concat(frames, ignore_index=True) if frames else _empty_long_frame()
    return long, all_warnings


def _parse_ifind_narrow_sheet(*, rows: list[tuple[object, ...]], path: Path) -> pd.DataFrame:
    name_row_index: int | None = None
    for row_index, row in enumerate(rows):
        first_cell = "" if not row else str(row[0]).strip()
        if first_cell == "指标名称" and len(row) >= 2 and row[1] is not None:
            name_row_index = row_index
            break
    if name_row_index is None:
        return _empty_long_frame()

    indicator_names = {
        col_index: str(value).strip()
        for col_index, value in enumerate(rows[name_row_index][1:], start=1)
        if value is not None and str(value).strip()
    }
    if not indicator_names:
        return _empty_long_frame()
    metadata = _metadata_by_column_from_narrow_rows(rows, name_row_index=name_row_index)
    formula = _first_formula(rows)
    data_start = name_row_index + 1
    rows_out: list[dict[str, object]] = []
    # iFinD 刷新后的 Excel 可能是一张多指标宽表；这里逐列展开为长表，
    # 后续再按指标名归入库存、仓单、进口等观察表。
    max_col_count = max((len(row) for row in rows), default=0)
    usable_columns = [
        col_index
        for col_index in range(1, max_col_count)
        if col_index in indicator_names
    ]
    for row in rows[data_start:]:
        trade_date = pd.to_datetime(row[0], errors="coerce")
        if pd.isna(trade_date):
            continue
        for col_index in usable_columns:
            if col_index >= len(row):
                continue
            indicator_name = indicator_names[col_index]
            value = pd.to_numeric(row[col_index], errors="coerce")
            if pd.isna(value):
                continue
            meta = metadata.get(col_index, {})
            rows_out.append(
                {
                    "trade_date": trade_date.date(),
                    "indicator_name": indicator_name,
                    "indicator_value": float(value),
                    "unit": meta.get("unit"),
                    "frequency": meta.get("frequency"),
                    "indicator_id": meta.get("indicator_id") or formula,
                    "source_name": meta.get("source_name")
                    or _source_for_narrow_indicator(indicator_name),
                    "update_time": meta.get("update_time"),
                    "source_file": str(path),
                }
            )
    if not rows_out:
        return _empty_long_frame()
    return pd.DataFrame(rows_out)


def _metadata_by_column_from_narrow_rows(
    rows: list[tuple[object, ...]],
    *,
    name_row_index: int,
) -> dict[int, dict[str, object]]:
    metadata: dict[int, dict[str, object]] = {}
    for row in rows[name_row_index + 1 : name_row_index + 7]:
        if not row:
            continue
        raw_key = str(row[0]).strip()
        normalized_key = IFIND_METADATA_KEYS.get(raw_key)
        if normalized_key is None:
            continue
        for col_index, value in enumerate(row[1:], start=1):
            if value is None or pd.isna(value):
                continue
            metadata.setdefault(col_index, {})[normalized_key] = value
    return metadata


def _metadata_value_from_narrow_rows(rows: list[tuple[object, ...]], key: str) -> object:
    for row in rows:
        if len(row) >= 2 and str(row[0]).strip() == key:
            return row[1]
    return None


def _first_formula(rows: list[tuple[object, ...]]) -> str | None:
    for row in rows:
        for cell in row:
            if isinstance(cell, str) and cell.strip().startswith("="):
                return cell.strip()
    return None


def _source_for_narrow_indicator(indicator_name: str) -> str:
    if "仓单" in indicator_name:
        return "郑州商品交易所/iFinD汇总"
    return "iFinD"


def _inventory_rows(long: pd.DataFrame) -> pd.DataFrame:
    if long.empty:
        return _empty_inventory_frame()
    subset = long.loc[long["indicator_name"].astype(str).str.contains("库存", na=False)].copy()
    if subset.empty:
        return _empty_inventory_frame()
    subset["product_code"] = PRODUCT_CODE
    subset["inventory_value"] = subset["indicator_value"]
    subset["data_quality_flag"] = "REVIEW_REQUIRED"
    subset["human_review_required"] = True
    subset["remark"] = "iFinD manual export; inventory source/unit require human review"
    columns = [
        "trade_date",
        "product_code",
        "indicator_name",
        "inventory_value",
        "unit",
        "source_name",
        "indicator_id",
        "update_time",
        "source_file",
        "data_quality_flag",
        "human_review_required",
        "remark",
    ]
    return subset[columns].sort_values(["trade_date", "indicator_name"]).reset_index(drop=True)


def _spot_rows(long: pd.DataFrame) -> pd.DataFrame:
    if long.empty:
        return _empty_spot_frame()
    pattern = "价格指数|现货|提货价|到厂价|期货收盘价"
    subset = long.loc[long["indicator_name"].astype(str).str.contains(pattern, na=False)].copy()
    if subset.empty:
        return _empty_spot_frame()
    subset["product_code"] = PRODUCT_CODE
    subset["data_quality_flag"] = "REVIEW_REQUIRED"
    subset["human_review_required"] = True
    subset["remark"] = "iFinD manual export; spot region/source require human review"
    columns = [
        "trade_date",
        "product_code",
        "indicator_name",
        "indicator_value",
        "unit",
        "source_name",
        "indicator_id",
        "update_time",
        "source_file",
        "data_quality_flag",
        "human_review_required",
        "remark",
    ]
    return subset[columns].sort_values(["trade_date", "indicator_name"]).reset_index(drop=True)


def _basis_rows(long: pd.DataFrame) -> pd.DataFrame:
    if long.empty:
        return _empty_basis_frame()
    # 基差先保留为观察口径：现货指数 vs iFinD 活跃合约，不能直接等同主力可交易合约。
    pivot = long.pivot_table(
        index="trade_date",
        columns="indicator_name",
        values="indicator_value",
        aggfunc="last",
    )
    spot_col = _first_column(pivot, ("中国棉花价格指数:3128B",))
    futures_col = _first_column(pivot, ("期货收盘价(活跃合约):棉花",))
    basis_col = _first_column(pivot, ("基差",))
    if spot_col is None or futures_col is None:
        return _empty_basis_frame()
    rows: list[dict[str, object]] = []
    metadata = long.drop_duplicates("indicator_name").set_index("indicator_name")
    for trade_date, row in pivot.iterrows():
        spot = _float_or_none(row.get(spot_col))
        futures = _float_or_none(row.get(futures_col))
        if spot is None or futures is None:
            continue
        basis = _float_or_none(row.get(basis_col)) if basis_col is not None else spot - futures
        source = _metadata_value(metadata, basis_col or spot_col, "source_name")
        rows.append(
            {
                "trade_date": trade_date,
                "product_code": PRODUCT_CODE,
                "region": "CCIndex_3128B_vs_iFinD_active_contract",
                "spot_price": spot,
                "futures_contract": "IFIND_ACTIVE_CONTRACT_REVIEW_REQUIRED",
                "futures_settle": futures,
                "basis": basis,
                "source_name": source or "iFinD",
                "spot_indicator_name": spot_col,
                "futures_indicator_name": futures_col,
                "basis_indicator_name": basis_col or "computed_spot_minus_futures",
                "unit": _metadata_value(metadata, spot_col, "unit") or "元/吨",
                "data_quality_flag": "REVIEW_REQUIRED",
                "human_review_required": True,
                "remark": (
                    "iFinD active-contract basis; exact futures contract mapping "
                    "requires human review"
                ),
            }
        )
    if not rows:
        return _empty_basis_frame()
    return pd.DataFrame(rows).sort_values("trade_date").reset_index(drop=True)


def _warehouse_receipt_rows(long: pd.DataFrame) -> pd.DataFrame:
    if long.empty:
        return _empty_warehouse_receipt_frame()
    mask = long["indicator_name"].astype(str).map(
        lambda value: "仓单" in value and "基差" not in value
    )
    subset = long.loc[mask].copy()
    if subset.empty:
        return _empty_warehouse_receipt_frame()
    subset["product_code"] = PRODUCT_CODE
    subset["warehouse_receipt"] = subset["indicator_value"]
    subset["data_quality_flag"] = "REVIEW_REQUIRED"
    subset["human_review_required"] = True
    subset["remark"] = "CZCE warehouse receipt series from iFinD/Wind aggregation"
    columns = [
        "trade_date",
        "product_code",
        "indicator_name",
        "warehouse_receipt",
        "unit",
        "source_name",
        "indicator_id",
        "update_time",
        "source_file",
        "data_quality_flag",
        "human_review_required",
        "remark",
    ]
    return subset[columns].sort_values(["trade_date", "indicator_name"]).reset_index(drop=True)


def _import_rows(long: pd.DataFrame) -> pd.DataFrame:
    if long.empty:
        return _empty_import_frame()
    subset = long.loc[long["indicator_name"].astype(str).str.contains("进口", na=False)].copy()
    if subset.empty:
        return _empty_import_frame()
    subset["product_code"] = PRODUCT_CODE
    subset["import_value"] = subset["indicator_value"]
    subset["data_quality_flag"] = "REVIEW_REQUIRED"
    subset["human_review_required"] = True
    # 进口数据多为月频，发布日期与统计期可能不等同于交易日，先作为解释观察保留。
    subset["remark"] = "iFinD manual export; import period/unit/release lag require human review"
    columns = [
        "trade_date",
        "product_code",
        "indicator_name",
        "import_value",
        "unit",
        "frequency",
        "source_name",
        "indicator_id",
        "update_time",
        "source_file",
        "data_quality_flag",
        "human_review_required",
        "remark",
    ]
    return subset[columns].sort_values(["trade_date", "indicator_name"]).reset_index(drop=True)


def _field_metadata_rows(
    long: pd.DataFrame,
    *,
    textile_chain: pd.DataFrame | None = None,
) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    if not long.empty:
        metadata = (
            long[
                [
                    "indicator_name",
                    "frequency",
                    "unit",
                    "indicator_id",
                    "source_name",
                    "update_time",
                    "source_file",
                ]
            ]
            .drop_duplicates()
            .copy()
        )
        # 元数据来自 iFinD 导出文件本身，用于审计字段来源，不代表已经形成交易信号。
        metadata["dataset_type"] = metadata["indicator_name"].astype(str).map(
            _dataset_type_for_indicator
        )
        metadata["raw_indicator_name"] = metadata["indicator_name"]
        metadata["product_code"] = PRODUCT_CODE
        metadata["metadata_status"] = "FROM_IFIND_EXPORT"
        metadata["human_review_required"] = True
        frames.append(metadata)
    if textile_chain is not None and not textile_chain.empty:
        textile_working = textile_chain.copy()
        if "raw_indicator_name" not in textile_working.columns:
            textile_working["raw_indicator_name"] = textile_working["indicator_name"]
        textile_metadata = (
            textile_working[
                [
                    "indicator_name",
                    "raw_indicator_name",
                    "metric_name",
                    "frequency",
                    "unit",
                    "indicator_id",
                    "source_name",
                    "source_file",
                ]
            ]
            .drop_duplicates()
            .copy()
        )
        textile_metadata["indicator_name"] = (
            textile_metadata["indicator_name"].astype(str)
            + ":"
            + textile_metadata["metric_name"].astype(str)
        )
        textile_metadata["raw_indicator_name"] = (
            textile_metadata["raw_indicator_name"].astype(str)
            + ":"
            + textile_metadata["metric_name"].astype(str)
        )
        textile_metadata["update_time"] = None
        textile_metadata["dataset_type"] = "textile_chain"
        textile_metadata["product_code"] = PRODUCT_CODE
        textile_metadata["metadata_status"] = "FROM_TTEB_BLOCK_WORKBOOK"
        textile_metadata["human_review_required"] = True
        frames.append(textile_metadata.drop(columns=["metric_name"]))
    if not frames:
        return _empty_field_metadata_frame()
    metadata = pd.concat(frames, ignore_index=True)
    columns = [
        "product_code",
        "dataset_type",
        "indicator_name",
        "raw_indicator_name",
        "frequency",
        "unit",
        "indicator_id",
        "source_name",
        "update_time",
        "source_file",
        "metadata_status",
        "human_review_required",
    ]
    return metadata[columns].sort_values(["dataset_type", "indicator_name"]).reset_index(drop=True)


def _dataset_type_for_indicator(indicator_name: str) -> str:
    if "仓单" in indicator_name and "基差" not in indicator_name:
        return "warehouse_receipt"
    if any(
        keyword in indicator_name
        for keyword in ("纺企", "织厂", "棉纱", "坯布", "负荷", "开工", "订单", "利润")
    ):
        return "textile_chain"
    if "库存" in indicator_name:
        return "inventory"
    if "基差" in indicator_name:
        return "basis"
    if "进口" in indicator_name:
        return "import"
    if any(keyword in indicator_name for keyword in ("纺织", "棉纱", "坯布", "订单", "利润")):
        return "textile_chain"
    if any(
        keyword in indicator_name
        for keyword in ("价格指数", "现货", "提货价", "到厂价", "期货收盘价")
    ):
        return "spot_price"
    return "other"


def _first_column(frame: pd.DataFrame, candidates: tuple[str, ...]) -> str | None:
    columns = [str(column) for column in frame.columns]
    for candidate in candidates:
        for column in columns:
            if column == candidate:
                return column
    return None


def _metadata_value(metadata: pd.DataFrame, indicator_name: str | None, column: str) -> object:
    if indicator_name is None or indicator_name not in metadata.index:
        return None
    value = metadata.loc[indicator_name, column]
    if isinstance(value, pd.Series):
        value = value.iloc[0]
    if pd.isna(value):
        return None
    return value


def _dataset_warnings(
    *,
    input_files: tuple[Path, ...],
    long: pd.DataFrame,
    basis: pd.DataFrame,
    textile_chain: pd.DataFrame,
) -> list[FundamentalObservationWarningRecord]:
    warnings: list[FundamentalObservationWarningRecord] = [
        FundamentalObservationWarningRecord(
            severity="INFO",
            warning_code="FUNDAMENTAL_SIGNAL_NOT_CONNECTED",
            message="R53 基本面观察不进入 signal matrix 或 composite_score。",
            affected_count=int(len(long) + len(textile_chain)),
            human_review_required=("fundamental_signal_rule_before_use",),
        )
    ]
    indicators = set() if long.empty else set(long["indicator_name"].astype(str))
    if not any("仓单" in item and "基差" not in item for item in indicators):
        warnings.append(
            FundamentalObservationWarningRecord(
                severity="WARN",
                warning_code="WAREHOUSE_RECEIPT_QUANTITY_MISSING",
                message="未识别到仓单数量序列；现货/基差文件不能替代仓单数量。",
                human_review_required=("warehouse_receipt_quantity_source",),
            )
        )
    if basis.empty:
        warnings.append(
            FundamentalObservationWarningRecord(
                severity="WARN",
                warning_code="BASIS_SERIES_NOT_READY",
                message="未能从现货价和期货活跃合约价格中构造基差序列。",
                human_review_required=("basis_spot_source_and_region",),
            )
        )
    if not any("进口" in path.name for path in input_files):
        warnings.append(
            FundamentalObservationWarningRecord(
                severity="WARN",
                warning_code="IMPORT_INPUT_MISSING",
                message="未发现进口数据文件。",
                human_review_required=("import_period_and_unit",),
            )
        )
    textile_file_exists = any(_looks_like_textile_chain_file(path) for path in input_files)
    if textile_chain.empty and not textile_file_exists:
        warnings.append(
            FundamentalObservationWarningRecord(
                severity="WARN",
                warning_code="TEXTILE_CHAIN_INPUT_MISSING",
                message="纺织链条数据暂缺：棉纱/坯布库存、开机率、订单、利润未接入。",
                human_review_required=("textile_chain_missing_data",),
            )
        )
    elif textile_chain.empty:
        warnings.append(
            FundamentalObservationWarningRecord(
                severity="WARN",
                warning_code="TEXTILE_CHAIN_INPUT_UNPARSED",
                message="已发现纺织链文件，但未能生成标准化纺织链观察表。",
                human_review_required=("official_fundamental_field_interpretation",),
            )
        )
    if not textile_chain.empty:
        stale_indicators = _stale_textile_indicators(textile_chain)
        if stale_indicators:
            warnings.append(
                FundamentalObservationWarningRecord(
                    severity="WARN",
                    warning_code="TEXTILE_CHAIN_FIELD_REVIEW_REQUIRED",
                    message=(
                        "部分 TTEB 纺织链指标未更新至纺织链最新日期，可能存在口径切换或停更："
                        + "；".join(
                            f"{name} 最后日期 {last_date}"
                            for name, last_date in stale_indicators.items()
                        )
                        + "。请人工确认是否存在替代指标。"
                    ),
                    affected_count=len(stale_indicators),
                    human_review_required=("official_fundamental_field_interpretation",),
                )
            )
    return warnings


def _stale_textile_indicators(textile_chain: pd.DataFrame) -> dict[str, str]:
    if textile_chain.empty:
        return {}
    working = textile_chain.copy()
    working["_trade_date"] = pd.to_datetime(working["trade_date"], errors="coerce")
    working = working.dropna(subset=["_trade_date"])
    if working.empty:
        return {}
    latest = working["_trade_date"].max().date()
    max_by_indicator = working.groupby("indicator_name")["_trade_date"].max().dt.date
    stale: dict[str, str] = {}
    for indicator_name, indicator_latest in max_by_indicator.items():
        # 纺织链是周频观察，超过两周未更新即进入人工复核清单，避免误把停更口径当作最新状态。
        if (latest - indicator_latest).days > 14:
            stale[str(indicator_name)] = indicator_latest.isoformat()
    return stale


def _dataset_summaries(
    *,
    inventory: pd.DataFrame,
    basis: pd.DataFrame,
    spot: pd.DataFrame,
    warehouse_receipt: pd.DataFrame,
    import_rows: pd.DataFrame,
    textile_chain: pd.DataFrame,
    input_files: tuple[Path, ...],
) -> list[FundamentalObservationDatasetSummary]:
    sources = tuple(str(path) for path in input_files)
    return [
        _summary_for("inventory", inventory, sources, "库存观察，人工复核单位和来源"),
        _summary_for("basis", basis, sources, "基差观察，活跃合约口径需人工复核"),
        _summary_for("spot_price", spot, sources, "现货/到厂/提货价格观察"),
        _summary_for(
            "warehouse_receipt",
            warehouse_receipt,
            sources,
            "郑商所仓单数量观察，iFinD/Wind 汇总口径需人工复核",
        ),
        _summary_for(
            "import",
            import_rows,
            sources,
            "进口数量/金额月频观察，统计期、单位和发布时间需人工复核",
        ),
        FundamentalObservationDatasetSummary(
            dataset_type="textile_chain",
            status=(
                "MISSING_INPUT" if textile_chain.empty else "READY_WITH_REVIEW"
            ),
            row_count=int(len(textile_chain)),
            date_start=(
                None
                if textile_chain.empty
                else pd.to_datetime(textile_chain["trade_date"]).min().date().isoformat()
            ),
            date_end=(
                None
                if textile_chain.empty
                else pd.to_datetime(textile_chain["trade_date"]).max().date().isoformat()
            ),
            source_files=sources,
            notes=(
                "TTEB 纱厂/织厂开工与库存观察，单位和口径需人工复核"
                if not textile_chain.empty
                else "用户确认暂时无法取得纺织链条数据"
            ),
        ),
    ]


def _summary_for(
    dataset_type: str,
    frame: pd.DataFrame,
    source_files: tuple[str, ...],
    notes: str,
) -> FundamentalObservationDatasetSummary:
    if frame.empty:
        return FundamentalObservationDatasetSummary(
            dataset_type=dataset_type,
            status="MISSING_INPUT",
            row_count=0,
            date_start=None,
            date_end=None,
            source_files=source_files,
            notes=notes,
        )
    dates = pd.to_datetime(frame["trade_date"])
    return FundamentalObservationDatasetSummary(
        dataset_type=dataset_type,
        status="READY_WITH_REVIEW",
        row_count=int(len(frame)),
        date_start=dates.min().date().isoformat(),
        date_end=dates.max().date().isoformat(),
        source_files=source_files,
        notes=notes,
    )


def _write_outputs(
    *,
    result: ResearchFundamentalObservationResult,
    inventory: pd.DataFrame,
    basis: pd.DataFrame,
    spot: pd.DataFrame,
    warehouse_receipt: pd.DataFrame,
    import_rows: pd.DataFrame,
    textile_chain: pd.DataFrame,
    field_metadata: pd.DataFrame,
) -> None:
    result.inventory_path.parent.mkdir(parents=True, exist_ok=True)
    _write_table(inventory, result.inventory_path)
    _write_table(basis, result.basis_path)
    _write_table(spot, result.spot_path)
    _write_table(warehouse_receipt, result.warehouse_receipt_path)
    _write_table(import_rows, result.import_path)
    _write_table(textile_chain, result.textile_chain_path)
    _write_field_metadata_csv(field_metadata, result.field_metadata_csv_path)
    _write_quality_csv(result)
    _write_warning_csv(result)
    _write_json(
        result=result,
        inventory=inventory,
        basis=basis,
        spot=spot,
        warehouse_receipt=warehouse_receipt,
        import_rows=import_rows,
        textile_chain=textile_chain,
        field_metadata=field_metadata,
    )
    _write_manifest(result)
    result.markdown_path.parent.mkdir(parents=True, exist_ok=True)
    result.markdown_path.write_text(
        _render_markdown(
            result=result,
            inventory=inventory,
            basis=basis,
            spot=spot,
            warehouse_receipt=warehouse_receipt,
            import_rows=import_rows,
            textile_chain=textile_chain,
            field_metadata=field_metadata,
        ),
        encoding="utf-8",
    )


def _write_table(frame: pd.DataFrame, path: Path) -> None:
    if path.suffix.lower() == ".csv":
        frame.to_csv(path, index=False, encoding="utf-8-sig")
        return
    frame.to_parquet(path, index=False)
    csv_path = path.with_suffix(".csv")
    frame.to_csv(csv_path, index=False, encoding="utf-8-sig")


def _write_quality_csv(result: ResearchFundamentalObservationResult) -> None:
    with result.quality_csv_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "dataset_type",
                "status",
                "row_count",
                "date_start",
                "date_end",
                "notes",
                "source_files",
            ],
        )
        writer.writeheader()
        for summary in result.dataset_summaries:
            writer.writerow(
                {
                    "dataset_type": summary.dataset_type,
                    "status": summary.status,
                    "row_count": summary.row_count,
                    "date_start": summary.date_start,
                    "date_end": summary.date_end,
                    "notes": summary.notes,
                    "source_files": ";".join(summary.source_files),
                }
            )


def _write_field_metadata_csv(frame: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    frame.to_csv(path, index=False, encoding="utf-8-sig")


def _write_warning_csv(result: ResearchFundamentalObservationResult) -> None:
    result.warning_csv_path.parent.mkdir(parents=True, exist_ok=True)
    with result.warning_csv_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "severity",
                "warning_code",
                "message",
                "affected_count",
                "human_review_required",
            ],
        )
        writer.writeheader()
        for warning in result.warning_records:
            writer.writerow(
                {
                    "severity": warning.severity,
                    "warning_code": warning.warning_code,
                    "message": warning.message,
                    "affected_count": warning.affected_count,
                    "human_review_required": ";".join(warning.human_review_required),
                }
            )


def _write_json(
    *,
    result: ResearchFundamentalObservationResult,
    inventory: pd.DataFrame,
    basis: pd.DataFrame,
    spot: pd.DataFrame,
    warehouse_receipt: pd.DataFrame,
    import_rows: pd.DataFrame,
    textile_chain: pd.DataFrame,
    field_metadata: pd.DataFrame,
) -> None:
    payload = {
        "report_type": "fundamental_observation",
        "rule_version": FUNDAMENTAL_OBSERVATION_VERSION,
        "generated_at": utc_now().isoformat(),
        "summary": result.to_summary(),
        "latest_observations": {
            "inventory": _latest_rows(inventory, key="indicator_name"),
            "basis": _latest_rows(basis, key="region"),
            "spot_price": _latest_rows(spot, key="indicator_name"),
            "warehouse_receipt": _latest_rows(warehouse_receipt, key="indicator_name"),
            "import": _latest_rows(import_rows, key="indicator_name"),
            "textile_chain": _latest_rows(textile_chain, key="indicator_name"),
        },
        "field_metadata_rows": field_metadata.to_dict(orient="records"),
        "fundamental_signal_status": "not_connected",
    }
    result.json_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True, default=str) + "\n",
        encoding="utf-8",
    )


def _write_manifest(result: ResearchFundamentalObservationResult) -> None:
    manifest = {
        "report_type": "fundamental_observation",
        "rule_version": FUNDAMENTAL_OBSERVATION_VERSION,
        "generated_at": utc_now().isoformat(),
        "run_id": result.run_id,
        "product_code": PRODUCT_CODE,
        "status": result.status,
        "data_asof": None if result.data_asof is None else result.data_asof.isoformat(),
        "source_dir": str(result.source_dir),
        "inventory_path": str(result.inventory_path),
        "basis_path": str(result.basis_path),
        "spot_path": str(result.spot_path),
        "warehouse_receipt_path": str(result.warehouse_receipt_path),
        "import_path": str(result.import_path),
        "textile_chain_path": str(result.textile_chain_path),
        "field_metadata_csv_path": str(result.field_metadata_csv_path),
        "quality_csv_path": str(result.quality_csv_path),
        "warning_csv_path": str(result.warning_csv_path),
        "json_path": str(result.json_path),
        "markdown_path": str(result.markdown_path),
        "fundamental_signal_status": "not_connected",
        "human_review_required": list(result.human_review_required),
    }
    result.manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def _render_markdown(
    *,
    result: ResearchFundamentalObservationResult,
    inventory: pd.DataFrame,
    basis: pd.DataFrame,
    spot: pd.DataFrame,
    warehouse_receipt: pd.DataFrame,
    import_rows: pd.DataFrame,
    textile_chain: pd.DataFrame,
    field_metadata: pd.DataFrame,
) -> str:
    lines = [
        f"# CF 基本面观察报告 R53 - {_date_text(result.data_asof)}",
        "",
        "## 数据状态",
        "",
        "- 报告类型：`fundamental_observation`",
        f"- 状态：`{result.status}`",
        f"- 数据截至：`{_date_text(result.data_asof)}`",
        f"- 输入文件数：`{result.input_file_count}`",
        "- 基本面信号状态：`not_connected`",
        "- 本报告只做人工复核观察，不进入 signal matrix 或 composite_score。",
        "",
        "## 数据集质量",
        "",
        "| 数据集 | 状态 | 行数 | 日期范围 | 说明 |",
        "| --- | --- | ---: | --- | --- |",
    ]
    for summary in result.dataset_summaries:
        date_range = "-"
        if summary.date_start and summary.date_end:
            date_range = f"{summary.date_start} 至 {summary.date_end}"
        lines.append(
            f"| `{summary.dataset_type}` | `{summary.status}` | {summary.row_count} | "
            f"{date_range} | {summary.notes} |"
        )
    lines.extend(["", "## 字段元数据", ""])
    lines.append(f"- 字段元数据表：`{result.field_metadata_csv_path}`")
    lines.append("- iFinD 导出文件已提供指标名、频率、单位、指标ID、来源和更新时间。")
    lines.append("- 这些元数据可用于审计和解释，但基本面仍未接入自动交易信号。")
    lines.extend(
        _markdown_rows(
            field_metadata.to_dict(orient="records"),
            columns=(
                "dataset_type",
                "indicator_name",
                "raw_indicator_name",
                "unit",
                "source_name",
                "frequency",
                "update_time",
            ),
        )
    )
    lines.extend(["", "## 最新基差观察", ""])
    lines.extend(
        _markdown_rows(
            _latest_rows(basis, key="region"),
            columns=("trade_date", "region", "spot_price", "futures_settle", "basis"),
        )
    )
    lines.extend(["", "## 最新仓单观察", ""])
    lines.extend(
        _markdown_rows(
            _latest_rows(warehouse_receipt, key="indicator_name"),
            columns=("trade_date", "indicator_name", "warehouse_receipt", "unit"),
        )
    )
    lines.extend(["", "## 最新进口观察", ""])
    lines.extend(
        _markdown_rows(
            _latest_rows(import_rows, key="indicator_name"),
            columns=("trade_date", "indicator_name", "import_value", "unit", "frequency"),
        )
    )
    lines.extend(["", "## 最新纺织链观察", ""])
    lines.extend(
        _markdown_rows(
            _latest_rows(textile_chain, key="indicator_name"),
            columns=(
                "trade_date",
                "indicator_name",
                "raw_indicator_name",
                "metric_name",
                "indicator_value",
                "unit",
                "frequency",
            ),
        )
    )
    lines.extend(["", "## 最新库存观察", ""])
    lines.extend(
        _markdown_rows(
            _latest_rows(inventory, key="indicator_name"),
            columns=("trade_date", "indicator_name", "inventory_value", "unit"),
        )
    )
    lines.extend(["", "## 最新现货价格观察", ""])
    lines.extend(
        _markdown_rows(
            _latest_rows(spot, key="indicator_name"),
            columns=("trade_date", "indicator_name", "indicator_value", "unit"),
        )
    )
    lines.extend(
        [
            "",
            "## 缺失与复核",
            "",
        ]
    )
    for warning in result.warning_records:
        if warning.severity == "INFO":
            continue
        lines.append(f"- `{warning.warning_code}`：{warning.message}")
    lines.extend(
        [
            "",
            "## 研究边界",
            "",
            "- R53 不生成 `fundamental_signal`。",
            "- 基差采用 iFinD 活跃合约口径，未映射到真实可交易主力合约前只能人工观察。",
            "- 仓单数量按郑商所口径处理，当前已接入 iFinD/Wind 汇总后的数量序列；"
            "来源和单位仍需人工复核。",
            "- 进口数据按 iFinD 月频统计期保留为观察项；发布时间、统计期和单位必须人工复核，"
            "不得直接作为交易信号。",
            "- TTEB 纺织链已作为观察输入接入时，仍只用于人工复核，不生成 `fundamental_signal`。",
            "- 缺失的进口、仓单数量、纺织订单、棉纱利润等数据不得估算填补。",
            "- 本报告不构成交易指令。",
        ]
    )
    return "\n".join(lines) + "\n"


def _latest_rows(frame: pd.DataFrame, *, key: str) -> list[dict[str, object]]:
    if frame.empty or key not in frame.columns:
        return []
    latest = pd.to_datetime(frame["trade_date"]).max().date()
    subset = frame.loc[pd.to_datetime(frame["trade_date"]).dt.date.eq(latest)].copy()
    sort_columns = [column for column in (key, "metric_name") if column in subset.columns]
    return subset.sort_values(sort_columns).to_dict(orient="records")


def _markdown_rows(rows: list[dict[str, object]], *, columns: tuple[str, ...]) -> list[str]:
    if not rows:
        return ["- 暂无可用数据。"]
    lines = [
        "| " + " | ".join(columns) + " |",
        "| " + " | ".join("---" for _ in columns) + " |",
    ]
    for row in rows[:12]:
        lines.append("| " + " | ".join(_cell(row.get(column)) for column in columns) + " |")
    return lines


def _cell(value: object) -> str:
    if value is None or pd.isna(value):
        return "-"
    if isinstance(value, float):
        return f"{value:.2f}"
    return str(value)


def _max_date(frames: tuple[pd.DataFrame, ...]) -> date | None:
    dates: list[date] = []
    for frame in frames:
        if not frame.empty and "trade_date" in frame.columns:
            dates.append(pd.to_datetime(frame["trade_date"]).max().date())
    return max(dates) if dates else None


def _date_text(value: date | None) -> str:
    return "NA" if value is None else value.isoformat()


def _float_or_none(value: object) -> float | None:
    if value is None or pd.isna(value):
        return None
    return float(value)


def _empty_long_frame() -> pd.DataFrame:
    return pd.DataFrame(
        columns=[
            "trade_date",
            "indicator_name",
            "indicator_value",
            "unit",
            "frequency",
            "indicator_id",
            "source_name",
            "update_time",
            "source_file",
        ]
    )


def _empty_field_metadata_frame() -> pd.DataFrame:
    return pd.DataFrame(
        columns=[
            "product_code",
            "dataset_type",
            "indicator_name",
            "raw_indicator_name",
            "frequency",
            "unit",
            "indicator_id",
            "source_name",
            "update_time",
            "source_file",
            "metadata_status",
            "human_review_required",
        ]
    )


def _empty_inventory_frame() -> pd.DataFrame:
    return pd.DataFrame(
        columns=[
            "trade_date",
            "product_code",
            "indicator_name",
            "inventory_value",
            "unit",
            "source_name",
            "indicator_id",
            "update_time",
            "source_file",
            "data_quality_flag",
            "human_review_required",
            "remark",
        ]
    )


def _empty_basis_frame() -> pd.DataFrame:
    return pd.DataFrame(
        columns=[
            "trade_date",
            "product_code",
            "region",
            "spot_price",
            "futures_contract",
            "futures_settle",
            "basis",
            "source_name",
            "spot_indicator_name",
            "futures_indicator_name",
            "basis_indicator_name",
            "unit",
            "data_quality_flag",
            "human_review_required",
            "remark",
        ]
    )


def _empty_warehouse_receipt_frame() -> pd.DataFrame:
    return pd.DataFrame(
        columns=[
            "trade_date",
            "product_code",
            "indicator_name",
            "warehouse_receipt",
            "unit",
            "source_name",
            "indicator_id",
            "update_time",
            "source_file",
            "data_quality_flag",
            "human_review_required",
            "remark",
        ]
    )


def _empty_import_frame() -> pd.DataFrame:
    return pd.DataFrame(
        columns=[
            "trade_date",
            "product_code",
            "indicator_name",
            "import_value",
            "unit",
            "frequency",
            "source_name",
            "indicator_id",
            "update_time",
            "source_file",
            "data_quality_flag",
            "human_review_required",
            "remark",
        ]
    )


def _empty_spot_frame() -> pd.DataFrame:
    return pd.DataFrame(
        columns=[
            "trade_date",
            "product_code",
            "indicator_name",
            "indicator_value",
            "unit",
            "source_name",
            "indicator_id",
            "update_time",
            "source_file",
            "data_quality_flag",
            "human_review_required",
            "remark",
        ]
    )


def _empty_textile_chain_frame() -> pd.DataFrame:
    return pd.DataFrame(
        columns=[
            "trade_date",
            "product_code",
            "indicator_name",
            "raw_indicator_name",
            "metric_name",
            "indicator_value",
            "unit",
            "frequency",
            "indicator_id",
            "source_name",
            "source_file",
            "source_sheet",
            "data_quality_flag",
            "human_review_required",
            "remark",
        ]
    )


def _inventory_path(output_dir: Path | None) -> Path:
    return _output_root(output_dir) / f"{PRODUCT_CODE}_fundamental_inventory_daily.parquet"


def _basis_path(output_dir: Path | None) -> Path:
    return _output_root(output_dir) / f"{PRODUCT_CODE}_fundamental_basis_daily.parquet"


def _spot_path(output_dir: Path | None) -> Path:
    return _output_root(output_dir) / f"{PRODUCT_CODE}_fundamental_spot_price_daily.parquet"


def _warehouse_receipt_path(output_dir: Path | None) -> Path:
    return _output_root(output_dir) / f"{PRODUCT_CODE}_fundamental_warehouse_receipt_daily.parquet"


def _import_path(output_dir: Path | None) -> Path:
    return _output_root(output_dir) / f"{PRODUCT_CODE}_fundamental_import_daily.parquet"


def _textile_chain_path(output_dir: Path | None) -> Path:
    return _output_root(output_dir) / f"{PRODUCT_CODE}_fundamental_textile_chain_daily.parquet"


def _field_metadata_csv_path(output_dir: Path | None) -> Path:
    return _output_root(output_dir) / f"{PRODUCT_CODE}_fundamental_field_metadata.csv"


def _quality_csv_path(output_dir: Path | None) -> Path:
    return _output_root(output_dir) / f"{PRODUCT_CODE}_fundamental_observation_quality.csv"


def _warning_csv_path(report_output_dir: Path | None) -> Path:
    return _report_root(report_output_dir) / f"{PRODUCT_CODE}_fundamental_observation_warnings.csv"


def _json_path(output_dir: Path | None) -> Path:
    return _output_root(output_dir) / f"{PRODUCT_CODE}_fundamental_observation.json"


def _manifest_path(output_dir: Path | None) -> Path:
    return _output_root(output_dir) / f"{PRODUCT_CODE}_fundamental_observation_manifest.json"


def _markdown_path(report_output_dir: Path | None) -> Path:
    return _report_root(report_output_dir) / f"{PRODUCT_CODE}_fundamental_observation.md"


def _output_root(output_dir: Path | None) -> Path:
    return output_dir or data_dir() / "research" / PRODUCT_CODE / OUTPUT_DIR


def _report_root(report_output_dir: Path | None) -> Path:
    return report_output_dir or reports_dir() / "research" / REPORT_DIR


def _default_run_id() -> str:
    return f"r53_fundamental_observation_{PRODUCT_CODE}_{uuid.uuid4().hex[:8]}"
