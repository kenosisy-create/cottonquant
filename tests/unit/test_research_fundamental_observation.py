from __future__ import annotations

import csv
import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from typer.testing import CliRunner

from cotton_factor.cli.main import app
from cotton_factor.research_workbench import build_cf_fundamental_observation


def test_build_cf_fundamental_observation_from_ifind_manual_exports(
    tmp_path: Path,
) -> None:
    source_dir = tmp_path / "incoming" / "CF" / "fundamentals" / "manual"
    _write_sample_fundamental_inputs(source_dir)

    result = build_cf_fundamental_observation(
        source_dir=source_dir,
        output_dir=tmp_path / "research" / "fundamentals",
        report_output_dir=tmp_path / "reports" / "fundamentals",
        run_id="r53_unit",
    )

    assert result.status == "OBSERVATION_READY_WITH_WARNINGS"
    assert result.passed is True
    assert result.data_asof is not None
    assert result.data_asof.isoformat() == "2026-07-03"
    assert result.to_summary()["fundamental_signal_status"] == "not_connected"

    warning_codes = {warning.warning_code for warning in result.warning_records}
    assert "IMPORT_INPUT_NOT_REFRESHED" in warning_codes
    assert "TEXTILE_CHAIN_INPUT_MISSING" in warning_codes
    assert "WAREHOUSE_RECEIPT_QUANTITY_MISSING" not in warning_codes

    inventory = pd.read_parquet(result.inventory_path)
    basis = pd.read_parquet(result.basis_path)
    spot = pd.read_parquet(result.spot_path)
    warehouse = pd.read_parquet(result.warehouse_receipt_path)
    assert len(inventory) == 2
    assert len(basis) == 2
    assert len(spot) == 6
    assert len(warehouse) == 2
    latest_basis = basis.loc[basis["trade_date"].astype(str).eq("2026-07-03"), "basis"]
    assert float(latest_basis.iloc[0]) == 1786.0
    latest_warehouse = warehouse.loc[
        warehouse["trade_date"].astype(str).eq("2026-07-03"),
        "warehouse_receipt",
    ]
    assert float(latest_warehouse.iloc[0]) == 10955.0
    assert set(inventory["data_quality_flag"]) == {"REVIEW_REQUIRED"}
    assert set(basis["human_review_required"]) == {True}
    assert set(warehouse["source_name"]) == {"郑州商品交易所/iFinD汇总"}

    quality_csv = result.quality_csv_path.read_text(encoding="utf-8")
    assert "warehouse_receipt" in quality_csv
    assert "MISSING_INPUT" in quality_csv
    assert result.inventory_path.with_suffix(".csv").exists()
    assert result.basis_path.with_suffix(".csv").exists()
    assert result.spot_path.with_suffix(".csv").exists()
    assert result.warehouse_receipt_path.with_suffix(".csv").exists()
    assert result.field_metadata_csv_path.exists()

    metadata = pd.read_csv(result.field_metadata_csv_path)
    assert {"indicator_name", "unit", "source_name", "indicator_id"} <= set(metadata.columns)
    assert "中国棉花价格指数:3128B" in set(metadata["indicator_name"])
    assert "元/吨" in set(metadata["unit"])

    payload = json.loads(result.json_path.read_text(encoding="utf-8"))
    assert payload["fundamental_signal_status"] == "not_connected"
    assert payload["latest_observations"]["basis"]
    assert payload["latest_observations"]["warehouse_receipt"]
    assert payload["field_metadata_rows"]

    manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
    assert manifest["status"] == "OBSERVATION_READY_WITH_WARNINGS"
    assert manifest["fundamental_signal_status"] == "not_connected"
    assert Path(manifest["warehouse_receipt_path"]).exists()
    assert Path(manifest["field_metadata_csv_path"]).exists()

    markdown = result.markdown_path.read_text(encoding="utf-8")
    assert "CF 基本面观察报告 R53" in markdown
    assert "基本面信号状态" in markdown
    assert "字段元数据" in markdown
    assert "最新仓单观察" in markdown
    assert "郑商所口径" in markdown
    assert "本报告不构成交易指令" in markdown
    assert "不进入 signal matrix 或 composite_score" in markdown


def test_build_cf_fundamental_observation_parses_tteb_textile_chain(
    tmp_path: Path,
) -> None:
    source_dir = tmp_path / "incoming" / "CF" / "fundamentals" / "manual"
    _write_sample_fundamental_inputs(source_dir)
    _write_tteb_textile_chain_workbooks(source_dir)

    result = build_cf_fundamental_observation(
        source_dir=source_dir,
        output_dir=tmp_path / "research" / "fundamentals",
        report_output_dir=tmp_path / "reports" / "fundamentals",
        run_id="r53_textile_unit",
    )

    warning_codes = {warning.warning_code for warning in result.warning_records}
    assert result.input_file_count == 6
    assert "TEXTILE_CHAIN_INPUT_MISSING" not in warning_codes
    assert "TEXTILE_CHAIN_INPUT_UNPARSED" not in warning_codes
    assert "FUNDAMENTAL_XLSX_PARSE_FAILED" not in warning_codes
    assert result.data_asof is not None
    assert result.data_asof.isoformat() == "2026-07-04"

    textile = pd.read_parquet(result.textile_chain_path)
    assert len(textile) == 36
    assert set(textile["source_name"]) == {"TTEB"}
    assert set(textile["unit"]) == {"%", "天"}
    assert set(textile["metric_name"]) == {"日均", "周均", "月均"}
    assert {
        "纯棉纱厂负荷",
        "全棉坯布负荷",
        "纺企棉花库存",
        "纺企棉纱库存",
        "织厂棉纱库存",
        "全棉坯布库存",
    } <= set(textile["indicator_name"])

    quality_csv = result.quality_csv_path.read_text(encoding="utf-8")
    assert "textile_chain" in quality_csv
    assert "READY_WITH_REVIEW" in quality_csv
    assert result.textile_chain_path.with_suffix(".csv").exists()

    payload = json.loads(result.json_path.read_text(encoding="utf-8"))
    assert payload["latest_observations"]["textile_chain"]
    assert payload["summary"]["textile_chain_path"] == str(result.textile_chain_path)

    manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
    assert Path(manifest["textile_chain_path"]).exists()

    markdown = result.markdown_path.read_text(encoding="utf-8")
    assert "最新纺织链观察" in markdown
    assert "TTEB 纺织链已作为观察输入接入" in markdown


def test_build_cf_fundamental_observation_maps_textile_inventory_continuation(
    tmp_path: Path,
) -> None:
    source_dir = tmp_path / "incoming" / "CF" / "fundamentals" / "manual"
    _write_sample_fundamental_inputs(source_dir)
    _write_tteb_inventory_continuation_workbook(
        source_dir / "TTEB纱线综合库存口径切换.xlsx"
    )

    result = build_cf_fundamental_observation(
        source_dir=source_dir,
        output_dir=tmp_path / "research" / "fundamentals",
        report_output_dir=tmp_path / "reports" / "fundamentals",
        run_id="r53_textile_continuation_unit",
    )

    warning_codes = {warning.warning_code for warning in result.warning_records}
    assert "TEXTILE_CHAIN_FIELD_REVIEW_REQUIRED" not in warning_codes

    textile = pd.read_parquet(result.textile_chain_path)
    yarn_inventory = textile.loc[
        textile["indicator_name"].eq("纺企棉纱库存")
        & textile["metric_name"].eq("周均")
    ]
    assert pd.to_datetime(yarn_inventory["trade_date"]).max().date().isoformat() == (
        "2026-07-03"
    )
    assert set(yarn_inventory["raw_indicator_name"]) == {
        "纺企棉纱库存",
        "纱线综合库存",
    }
    assert "纱线综合库存" not in set(textile["indicator_name"])
    assert yarn_inventory["remark"].str.contains("按纺企棉纱库存延续口径处理").any()


def test_cli_build_cf_fundamental_observation(tmp_path: Path) -> None:
    source_dir = tmp_path / "incoming" / "CF" / "fundamentals" / "manual"
    _write_sample_fundamental_inputs(source_dir)

    result = CliRunner().invoke(
        app,
        [
            "research",
            "build-cf-fundamental-observation",
            "--source-dir",
            str(source_dir),
            "--output-dir",
            str(tmp_path / "research" / "fundamentals"),
            "--report-output-dir",
            str(tmp_path / "reports" / "fundamentals"),
            "--run-id",
            "r53_cli",
        ],
    )

    assert result.exit_code == 0, result.output
    output = json.loads(result.output)
    assert output["run_id"] == "r53_cli"
    assert output["status"] == "OBSERVATION_READY_WITH_WARNINGS"
    assert output["fundamental_signal_status"] == "not_connected"
    assert output["warning_count"] == 2
    assert Path(output["inventory_path"]).exists()
    assert Path(output["basis_path"]).exists()
    assert Path(output["spot_path"]).exists()
    assert Path(output["warehouse_receipt_path"]).exists()
    assert Path(output["textile_chain_path"]).exists()
    assert Path(output["field_metadata_csv_path"]).exists()
    assert Path(output["markdown_path"]).exists()


def test_build_cf_fundamental_observation_parses_refreshed_import_workbook(
    tmp_path: Path,
) -> None:
    source_dir = tmp_path / "incoming" / "CF" / "fundamentals" / "manual"
    _write_sample_fundamental_inputs(source_dir)
    (source_dir / "中国_进口数量_棉花.xlsx").unlink()
    _write_import_wide_workbook(source_dir / "中国_进口数量_棉花1.xlsx")

    result = build_cf_fundamental_observation(
        source_dir=source_dir,
        output_dir=tmp_path / "research" / "fundamentals",
        report_output_dir=tmp_path / "reports" / "fundamentals",
        run_id="r53_import_unit",
    )

    warning_codes = {warning.warning_code for warning in result.warning_records}
    assert "IMPORT_INPUT_NOT_REFRESHED" not in warning_codes
    assert result.to_summary()["import_path"] == str(result.import_path)

    import_frame = pd.read_parquet(result.import_path)
    assert len(import_frame) == 16
    assert set(import_frame["data_quality_flag"]) == {"REVIEW_REQUIRED"}
    assert set(import_frame["human_review_required"]) == {True}
    assert {
        "棉花:进口数量:当月值",
        "棉花:进口数量:累计值",
        "棉花:进口金额:当月值",
    } <= set(import_frame["indicator_name"])
    latest_quantity = import_frame.loc[
        import_frame["indicator_name"].eq("棉花:进口数量:当月值")
        & import_frame["trade_date"].astype(str).eq("2026-05-31"),
        "import_value",
    ]
    assert float(latest_quantity.iloc[0]) == 11.0

    payload = json.loads(result.json_path.read_text(encoding="utf-8"))
    assert payload["latest_observations"]["import"]
    import_summary = next(
        summary
        for summary in payload["summary"]["dataset_summaries"]
        if summary["dataset_type"] == "import"
    )
    assert import_summary["status"] == "READY_WITH_REVIEW"
    assert import_summary["row_count"] == 16

    markdown = result.markdown_path.read_text(encoding="utf-8")
    assert "最新进口观察" in markdown
    assert "进口数据按 iFinD 月频统计期保留为观察项" in markdown


def _write_sample_fundamental_inputs(source_dir: Path) -> None:
    source_dir.mkdir(parents=True, exist_ok=True)
    _write_ifind_wide_csv(
        source_dir / "库存数据_棉花.csv",
        {
            "中国:商业库存量:棉花": ["341.36", "374.53"],
        },
        dates=("2026-06-15", "2026-05-31"),
        units=("万吨",),
        sources=("中国棉花信息网",),
    )
    _write_ifind_wide_csv(
        source_dir / "现货指数、仓单基差、现货价、到场价数据0705.csv",
        {
            "中国棉花价格指数:3128B": ["17786", "17777"],
            "期货收盘价(活跃合约):棉花": ["16000", "15990"],
            "基差": ["1786", "1787"],
            "中国:现货平均价:棉花": ["17773.85", "17731.54"],
        },
        dates=("2026-07-03", "2026-07-02"),
        units=("元/吨", "元/吨", "元/吨", "元/吨"),
        sources=("全国棉花交易市场", "郑商所/iFinD", "iFinD", "国家统计"),
    )
    _write_warehouse_receipt_workbook(source_dir / "仓单数量一号棉.xlsx")
    _write_formula_only_workbook(source_dir / "中国_进口数量_棉花.xlsx")


def _write_ifind_wide_csv(
    path: Path,
    series: dict[str, list[str]],
    *,
    dates: tuple[str, ...],
    units: tuple[str, ...],
    sources: tuple[str, ...],
) -> None:
    indicator_names = tuple(series)
    rows = [
        ("指标名称", *indicator_names),
        ("频率", *("日" for _ in indicator_names)),
        ("单位", *units),
        ("指标ID", *(f"S{i:04d}" for i, _ in enumerate(indicator_names, start=1))),
        ("来源", *sources),
        ("更新时间", *("2026-07-05" for _ in indicator_names)),
    ]
    for date_value_index, date_value in enumerate(dates):
        rows.append(
            (
                date_value,
                *(series[indicator_name][date_value_index] for indicator_name in indicator_names),
            )
        )
    with path.open("w", encoding="gbk", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerows(rows)


def _write_formula_only_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "中国_进口数量_棉花_当月值"
    sheet["A1"] = "==edb()"
    workbook.save(path)


def _write_import_wide_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "中国_进口数量_棉花_已刷新"
    sheet.append(["=[1]!HX_IFIND_EDB(0)"])
    sheet.append(
        [
            "指标名称",
            "棉花:进口金额:当月同比",
            "棉花:进口金额:当月值",
            "棉花:进口金额:累计同比",
            "棉花:进口金额:累计值",
            "棉花:进口数量:当月同比",
            "棉花:进口数量:当月值",
            "棉花:进口数量:累计同比",
            "棉花:进口数量:累计值",
        ]
    )
    sheet.append(["频率", *("月" for _ in range(8))])
    sheet.append(["单位", "%", "千美元", "%", "千美元", "%", "万吨", "%", "万吨"])
    sheet.append(["2026-05-31", 180.7, 191711, 69.7, 1380683, 214.7, 11, 89.9, 83])
    sheet.append(["2026-04-30", 134.1, 273605, 59.6, 1188991, 171.0, 17, 78.7, 71])
    workbook.save(path)


def _write_warehouse_receipt_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = "=[1]!HX_IFIND_EDB(0)"
    sheet["A2"] = "指标名称"
    sheet["B2"] = "仓单数量:一号棉"
    sheet["A3"] = "频率"
    sheet["B3"] = "日"
    sheet["A4"] = "单位"
    sheet["B4"] = "张"
    sheet["A5"] = "2026-07-03"
    sheet["B5"] = 10955
    sheet["A6"] = "2026-07-02"
    sheet["B6"] = 10999
    workbook.save(path)


def _write_tteb_textile_chain_workbooks(source_dir: Path) -> None:
    _write_tteb_load_workbook(source_dir / "TTEB纯棉布及纱开工率负荷数据.xlsx")
    _write_tteb_inventory_workbook(
        source_dir / "TTEB纱厂原料成品、布厂原料成品周度库存数据.xlsx"
    )
    (source_dir / "~$TTEB纱厂原料成品、布厂原料成品周度库存数据.xlsx").write_bytes(
        b"excel-lock-file"
    )


def _write_tteb_load_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "13.TTEB开机负荷%202512"
    sheet.append(
        [
            "产品",
            "日期",
            "日均",
            "周均",
            "月均",
            "",
            "产品",
            "日期",
            "日均",
            "周均",
            "月均",
        ]
    )
    sheet.append(
        [
            "纯棉纱厂负荷",
            "2026-06-27",
            55.1,
            55.0,
            54.8,
            "",
            "全棉坯布负荷",
            "2026-06-27",
            51.2,
            51.1,
            50.8,
        ]
    )
    sheet.append(
        [
            "纯棉纱厂负荷",
            "2026-07-04",
            56.4,
            56.0,
            55.5,
            "",
            "全棉坯布负荷",
            "2026-07-04",
            52.3,
            52.0,
            51.4,
        ]
    )
    workbook.save(path)


def _write_tteb_inventory_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "12-TTEB产业链库存（天数）202512"
    sheet.append(
        [
            "",
            "",
            "",
            "纺企棉花库存",
            "日期",
            "日均",
            "周均",
            "月均",
            "",
            "纺企棉纱库存",
            "日期",
            "日均",
            "周均",
            "月均",
            "",
            "织厂棉纱库存",
            "日期",
            "日均",
            "周均",
            "月均",
            "",
            "全棉坯布库存",
            "日期",
            "日均",
            "周均",
            "月均",
        ]
    )
    for trade_date, values in (
        ("2026-06-27", (28.0, 24.0, 9.0, 31.0)),
        ("2026-07-04", (29.0, 25.0, 10.0, 32.0)),
    ):
        sheet.append(
            [
                "",
                "",
                "",
                "纺企棉花库存",
                trade_date,
                values[0],
                values[0] + 0.1,
                values[0] + 0.2,
                "",
                "纺企棉纱库存",
                trade_date,
                values[1],
                values[1] + 0.1,
                values[1] + 0.2,
                "",
                "织厂棉纱库存",
                trade_date,
                values[2],
                values[2] + 0.1,
                values[2] + 0.2,
                "",
                "全棉坯布库存",
                trade_date,
                values[3],
                values[3] + 0.1,
                values[3] + 0.2,
            ]
        )
    workbook.save(path)


def _write_tteb_inventory_continuation_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TTEB库存口径切换"
    sheet.append(["产品", "日期", "日均", "周均", "月均"])
    sheet.append(["纺企棉纱库存", "2025-12-12", 24.0, 24.1, 24.2])
    sheet.append(["纱线综合库存", "2025-12-19", 24.3, 24.4, 24.5])
    sheet.append(["纱线综合库存", "2026-07-03", 25.0, 25.1, 25.2])
    workbook.save(path)
