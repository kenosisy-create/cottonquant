"""Create the initialization registries from locked, page-audited metadata.

This script never edits source PDFs. It verifies source/target copies first and
refuses to overwrite an existing registry, preserving the initialization audit.
"""

from __future__ import annotations

import csv
import hashlib
import json
import sys
import unicodedata
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
SOURCE_ROOT = Path(r"D:\researchreports\bond")
RAW_REPORTS = ROOT / "00_inbox" / "raw_reports"
PENDING_REVIEW = ROOT / "00_inbox" / "pending_review"
INGESTED_AT = datetime.now(ZoneInfo("Asia/Shanghai")).isoformat(timespec="seconds")
RUN_ID = "init_20260815_registry_v1"


REPORT_COLUMNS = [
    "report_id", "relative_path", "source_original_path", "filename", "sha256",
    "file_size_bytes", "page_count", "title", "institution", "co_brand",
    "authors_display", "publish_date", "publish_date_text", "publish_date_precision",
    "filename_date", "data_cutoff", "latest_explicit_data_date", "data_cutoff_scope",
    "primary_topic", "topic_tags", "report_type", "research_horizon",
    "instrument_scope", "has_text_layer", "text_page_coverage_pct", "text_char_count",
    "ocr_requirement", "visual_review_requirement", "chart_density", "duplicate_status",
    "duplicate_of_report_id", "partial_overlap_group", "extraction_quality",
    "manual_review_status", "ingested_at", "notes",
]

EVIDENCE_COLUMNS = [
    "evidence_id", "report_id", "field_name", "field_value", "evidence_class",
    "source_type", "page_start", "page_end", "evidence_data_cutoff", "cutoff_scope",
    "extraction_method", "confidence", "review_status", "note",
]

PROCESSING_COLUMNS = [
    "run_id", "step_no", "timestamp", "object_type", "object_id", "stage", "action",
    "tool", "tool_version", "input_sha256", "output_relative_path", "output_sha256",
    "status", "records_or_pages", "config_sha256", "git_commit", "message",
]

ISSUE_COLUMNS = [
    "issue_id", "report_id", "asset_id", "field_name", "page_start", "page_end",
    "issue_type", "severity", "evidence_class", "description", "status", "resolution",
    "detected_at", "reviewed_at",
]

ASSET_COLUMNS = [
    "asset_id", "relative_path", "source_original_path", "filename", "file_type",
    "sha256", "file_size_bytes", "asset_role", "associated_report_id", "institution",
    "asset_date", "sheet_count", "extraction_quality", "manual_review_status", "notes",
]


REPORTS = [
    {
        "filename": "【期债半年报】水活则鱼动，把握流动性改善的机遇.pdf",
        "sha256": "c406ce66269f6c16627ef25fa6806e5f9060549bd1700fb9f4e3b12d118b7db3",
        "page_count": 15,
        "title": "2026年半年度报告：水活则鱼动，把握流动性改善的机遇",
        "institution": "中信建投期货有限公司", "co_brand": "",
        "authors": [("孙玉龙", "作者姓名", 1)],
        "publish_date": "2026-06-30", "publish_date_text": "2026 年6 月 30 日",
        "publish_date_precision": "day", "filename_date": "",
        "latest_date": "", "cutoff_scope": "none", "cutoff_pages": "",
        "primary_topic": "国债期货流动性与期现策略",
        "topic_tags": "资金面|跨期价差|基差|IRR|Carry|流动性",
        "report_type": "futures_microstructure", "research_horizon": "2026H2",
        "instrument_scope": "TS|TF|T|TL|跨期价差|基差|IRR|Carry",
        "chart_density": "high", "extraction_quality": "medium",
        "overlap_group": "", "institution_page": 14, "author_page": 1,
        "notes": "[U] 未声明全篇统一数据截止日；PDF属性标题为模板名、作者为“1”，以封面/正文为准。",
    },
    {
        "filename": "20260706-华泰期货-国债半年报：反弹之后，等待拐点.pdf",
        "sha256": "c5fb1a641cf93b7ca6fa7176c8e1a85295343788e8fe3346138a463d393e9ad7",
        "page_count": 14,
        "title": "国债：反弹之后，等待拐点",
        "institution": "华泰期货有限公司", "co_brand": "华泰期货研究院",
        "authors": [("徐闻宇", "本期分析研究员", 1)],
        "publish_date": "2026-07-06", "publish_date_text": "2026 年 07 月 06 日",
        "publish_date_precision": "day", "filename_date": "2026-07-06",
        "latest_date": "2026-06-30", "cutoff_scope": "section_specific", "cutoff_pages": "4",
        "primary_topic": "2026年下半年国债与利率展望",
        "topic_tags": "宏观基本面|货币政策|供给|流动性|收益率曲线|期债",
        "report_type": "half_year_outlook", "research_horizon": "2026H2",
        "instrument_scope": "现券|收益率曲线|国债期货",
        "chart_density": "high", "extraction_quality": "high",
        "overlap_group": "overlap_huatai_futures_2026h2", "institution_page": 13,
        "notes": "[U] 2026-06-30仅为局部最新明确日期，不代表全篇统一截止日。",
    },
    {
        "filename": "20260706-建信期货-国债半年报：1.65％—1.95％，利率的窄走廊与新均衡.pdf",
        "sha256": "ffa011069cee4a472c5dff0c0471cb88e44cd7e1e42ae483fed777326d3246a8",
        "page_count": 19,
        "title": "1.65%—1.95%：利率的窄走廊与新均衡",
        "institution": "建信期货有限责任公司", "co_brand": "研究发展部",
        "authors": [("何卓乔", "研究员（宏观贵金属）", 1), ("黄雯昕", "研究员（国债集运）", 1), ("聂嘉怡", "研究员（股指）", 1)],
        "publish_date": "2026-07-06", "publish_date_text": "2026 年 7 月 6 日",
        "publish_date_precision": "day", "filename_date": "2026-07-06",
        "latest_date": "2026-06-26", "cutoff_scope": "mixed", "cutoff_pages": "5",
        "primary_topic": "利率走廊与下半年国债展望",
        "topic_tags": "货币政策|利率走廊|收益率曲线|期限利差|期债",
        "report_type": "half_year_outlook", "research_horizon": "2026H2",
        "instrument_scope": "现券|收益率曲线|期限利差|国债期货",
        "chart_density": "high", "extraction_quality": "medium",
        "overlap_group": "", "institution_page": 19,
        "notes": "[U] 市场快照、历史样本和预测表存在多口径日期。",
    },
    {
        "filename": "20260726-国泰海通证券-债券基金周度数据观察：30年国债ETF为何越涨越赎.pdf",
        "sha256": "97d85a7c1574f3397077083425af215cfe5ab9e95d4b1161309db7ffbb17eefe",
        "page_count": 11, "title": "30年国债ETF为何越涨越赎",
        "institution": "国泰海通证券股份有限公司", "co_brand": "研究所",
        "authors": [("唐元懋", "分析师", 1), ("熊航飞", "研究助理", 1)],
        "publish_date": "2026-07-26", "publish_date_text": "2026.07.26",
        "publish_date_precision": "day", "filename_date": "2026-07-26",
        "latest_date": "2026-07-24", "cutoff_scope": "mixed", "cutoff_pages": "4-6",
        "primary_topic": "30年国债ETF申赎与机构行为",
        "topic_tags": "债券基金|30年国债ETF|申赎|机构行为|CTD",
        "report_type": "fund_data", "research_horizon": "weekly",
        "instrument_scope": "30年国债ETF|现券|收益率曲线|TL|CTD",
        "chart_density": "high", "extraction_quality": "medium",
        "overlap_group": "overlap_fund_duration_2026q3", "institution_page": 11,
        "notes": "[U] ETF、现券、期货及基金持仓数据更新时间不同。",
    },
    {
        "filename": "20260727-国泰海通证券-市场策略周报：如何识别国债期货盘中的“量化”博弈特征.pdf",
        "sha256": "b793155f619a8ad86d9fb1f8eb5a9521d3c40c7dbb3f9cb8f342092d6d94b384",
        "page_count": 9, "title": "如何识别国债期货盘中的“量化”博弈特征",
        "institution": "国泰海通证券股份有限公司", "co_brand": "研究所",
        "authors": [("唐元懋", "分析师", 1), ("孙越", "分析师", 1)],
        "publish_date": "2026-07-27", "publish_date_text": "2026.07.27",
        "publish_date_precision": "day", "filename_date": "2026-07-27",
        "latest_date": "2026-07-24", "cutoff_scope": "section_specific", "cutoff_pages": "6-7",
        "primary_topic": "国债期货盘中交易结构识别",
        "topic_tags": "国债期货|成交持仓|CTD|基差|IRR|交易行为",
        "report_type": "futures_microstructure", "research_horizon": "weekly",
        "instrument_scope": "国债期货|成交持仓|CTD|基差|IRR",
        "chart_density": "high", "extraction_quality": "medium", "overlap_group": "",
        "institution_page": 9,
        "notes": "[U] “类量化”是报告构造的识别框架，不能当作账户身份事实。",
    },
    {
        "filename": "20260730-国泰海通证券-7月FOMC：沃什说了什么，票委做了什么.pdf",
        "sha256": "737662a2d6e52aed7d3fa01ad4e2879baa3e67a0747a02b74d17e5e520d32ca2",
        "page_count": 8, "title": "7月FOMC：沃什说了什么，票委做了什么",
        "institution": "国泰海通证券股份有限公司", "co_brand": "研究所",
        "authors": [("王一凡", "分析师", 1), ("唐元懋", "分析师", 1)],
        "publish_date": "2026-07-30", "publish_date_text": "2026.07.30",
        "publish_date_precision": "day", "filename_date": "2026-07-30",
        "latest_date": "", "cutoff_scope": "none", "cutoff_pages": "",
        "primary_topic": "7月FOMC事件与美债定价",
        "topic_tags": "FOMC|美联储|票委|美债曲线|事件点评",
        "report_type": "event_commentary", "research_horizon": "event_window",
        "instrument_scope": "美债现券|美债收益率曲线|美元",
        "chart_density": "medium", "extraction_quality": "high", "overlap_group": "",
        "institution_page": 8,
        "notes": "[U] 会议事件、盘中概率与历史数据库更新月份不同，未发现全篇统一截止日。",
    },
    {
        "filename": "20260802-国泰海通证券-透视股债跷跷板：细分资产相关性的K型分化.pdf",
        "sha256": "cfe3ee74ed15137041f3e4040e9de7b670365bb06b98de6a2127785a66cc9fac",
        "page_count": 10, "title": "透视股债跷跷板：细分资产相关性的K型分化",
        "institution": "国泰海通证券股份有限公司", "co_brand": "研究所",
        "authors": [("唐元懋", "分析师", 1), ("杜润琛", "分析师", 1)],
        "publish_date": "2026-08-02", "publish_date_text": "2026.08.02",
        "publish_date_precision": "day", "filename_date": "2026-08-02",
        "latest_date": "2026-07-31", "cutoff_scope": "mixed", "cutoff_pages": "6",
        "primary_topic": "股债相关性与细分资产分化",
        "topic_tags": "股债跷跷板|相关性|收益率曲线|跨资产",
        "report_type": "cross_asset_macro", "research_horizon": "weekly",
        "instrument_scope": "现券|收益率曲线|股票|跨资产",
        "chart_density": "high", "extraction_quality": "medium", "overlap_group": "",
        "institution_page": 10,
        "notes": "[U] 半年相关性样本与周度市场快照混合；不得将相关性升级为因果。",
    },
    {
        "filename": "20260804-国泰海通证券-机构行为周度跟踪：30年国债走强行情，会起波动吗，机构行为关注三个关键信号.pdf",
        "sha256": "05386acea6130991db6e0f662dbfcc67a1f00ed6fd6650eb2d594201f0d843b2",
        "page_count": 9, "title": "30年国债走强行情，会起波动吗：机构行为关注三个关键信号",
        "institution": "国泰海通证券股份有限公司", "co_brand": "研究所",
        "authors": [("唐元懋", "分析师", 1), ("汤志宇", "分析师", 1)],
        "publish_date": "2026-08-04", "publish_date_text": "2026.08.04",
        "publish_date_precision": "day", "filename_date": "2026-08-04",
        "latest_date": "2026-07-31", "cutoff_scope": "section_specific", "cutoff_pages": "3-4",
        "primary_topic": "30年国债机构行为信号",
        "topic_tags": "30年国债|机构行为|杠杆|一级市场|波动",
        "report_type": "institution_behavior", "research_horizon": "weekly",
        "instrument_scope": "30年国债现券|收益率曲线|机构行为",
        "chart_density": "high", "extraction_quality": "medium", "overlap_group": "",
        "institution_page": 9,
        "notes": "[U] 交易、杠杆和一级市场数据可能存在各自更新时差。",
    },
    {
        "filename": "20260809-国泰海通证券-30年活跃券如何切换：历史规律与26特6的“不稳定接棒”.pdf",
        "sha256": "c811a6f1f2c782709d9963acb6275d47a78a693dc3e6501145d04734df41365a",
        "page_count": 12, "title": "30年活跃券如何切换：历史规律与26特6的“不稳定接棒”",
        "institution": "国泰海通证券股份有限公司", "co_brand": "研究所",
        "authors": [("唐元懋", "分析师", 1), ("孙越", "分析师", 1)],
        "publish_date": "2026-08-09", "publish_date_text": "2026.08.09",
        "publish_date_precision": "day", "filename_date": "2026-08-09",
        "latest_date": "2026-08-07", "cutoff_scope": "mixed", "cutoff_pages": "8",
        "primary_topic": "30年国债活跃券切换与流动性",
        "topic_tags": "活跃券|30年国债|流动性|换券|供给",
        "report_type": "futures_microstructure", "research_horizon": "weekly",
        "instrument_scope": "30年国债现券|收益率曲线|流动性",
        "chart_density": "high", "extraction_quality": "medium", "overlap_group": "",
        "institution_page": 12,
        "notes": "[U] 历史切券样本、供给计划和当周数据口径混合。",
    },
    {
        "filename": "20260809-国泰海通证券-港交所国债期货上市首周：定价中枢与跨境联动.pdf",
        "sha256": "9fa333483e94e1694987ac258fb5c052f4fe06d022059392dcf1a37aeecb5095",
        "page_count": 10, "title": "港交所国债期货上市首周：定价中枢与跨境联动",
        "institution": "国泰海通证券股份有限公司", "co_brand": "研究所",
        "authors": [("唐元懋", "分析师", 1), ("孙越", "分析师", 1)],
        "publish_date": "2026-08-09", "publish_date_text": "2026.08.09",
        "publish_date_precision": "day", "filename_date": "2026-08-09",
        "latest_date": "2026-08-07", "cutoff_scope": "mixed", "cutoff_pages": "8",
        "primary_topic": "港交所国债期货定价与跨境联动",
        "topic_tags": "港交所国债期货|中金所|跨境联动|现金结算|期现",
        "report_type": "futures_microstructure", "research_horizon": "listing_week",
        "instrument_scope": "港交所国债期货|中金所国债期货|跨境联动",
        "chart_density": "high", "extraction_quality": "medium", "overlap_group": "",
        "institution_page": 10,
        "notes": "[U] 港交所现金结算与中金所实物交割不可混写为同一基差或收敛套利。",
    },
    {
        "filename": "20260809-华泰期货-国债周报：国债期货偏弱震荡，长端跌幅大于短端.pdf",
        "sha256": "5856700aced049fca482f870ab0be5fe4c2a964b9b5ecdc1ef21e798326ef93c",
        "page_count": 13, "title": "国债期货偏弱震荡，长端跌幅大于短端",
        "institution": "华泰期货有限公司", "co_brand": "华泰期货研究院",
        "authors": [("徐闻宇", "本期分析研究员", 12)],
        "publish_date": "2026-08-09", "publish_date_text": "2026-08-09",
        "publish_date_precision": "day", "filename_date": "2026-08-09",
        "latest_date": "2026-08-07", "cutoff_scope": "section_specific", "cutoff_pages": "1",
        "primary_topic": "国债期货周度市场与期现指标",
        "topic_tags": "TS|TF|T|TL|基差|净基差|IRR|Carry|跨期价差",
        "report_type": "weekly_market", "research_horizon": "weekly",
        "instrument_scope": "TS|TF|T|TL|跨期价差|基差|净基差|IRR|Carry",
        "chart_density": "high", "extraction_quality": "medium",
        "overlap_group": "overlap_huatai_futures_2026h2", "institution_page": 12,
        "notes": "[U] 2026-08-07为局部央行操作和期货收盘口径；月度宏观数据更早。沈嘉奇仅为联系人，未录为作者。",
    },
    {
        "filename": "20260810-国泰海通证券-债基久期高位之后：关注绩优基金久期“死叉”.pdf",
        "sha256": "feca128af69a3d3b762663926069bf4d99dcf8ad540c8320fa1cf465a5c11f70",
        "page_count": 9, "title": "债基久期高位之后：关注绩优基金久期“死叉”",
        "institution": "国泰海通证券股份有限公司", "co_brand": "研究所",
        "authors": [("唐元懋", "分析师", 1), ("熊航飞", "研究助理", 1)],
        "publish_date": "2026-08-10", "publish_date_text": "2026.08.10",
        "publish_date_precision": "day", "filename_date": "2026-08-10",
        "latest_date": "2026-08-08", "cutoff_scope": "mixed", "cutoff_pages": "5",
        "primary_topic": "债券基金久期与绩优基金行为",
        "topic_tags": "债券基金|久期|绩优基金|机构行为|持仓",
        "report_type": "institution_behavior", "research_horizon": "weekly",
        "instrument_scope": "债券基金|现券|收益率曲线",
        "chart_density": "high", "extraction_quality": "medium",
        "overlap_group": "overlap_fund_duration_2026q3", "institution_page": 9,
        "notes": "[U] 日度基金数据、周区间与2026-06-30持仓报告期同时存在。",
    },
    {
        "filename": "20260814-国泰君安期货-2026H2海外宏观经济及大类资产展望：旧序余寒，新质去伪存真.pdf",
        "sha256": "bb2de5c3b82fc0e29dc122c06d0ea37d069edfbf68c9718be01d9db511ff4294",
        "page_count": 38, "title": "旧序余寒，新质去伪存真——2026H2海外宏观经济及大类资产展望",
        "institution": "国泰君安期货有限公司", "co_brand": "",
        "authors": [("戴璐", "首席分析师、宏观总量组行政负责人", 1)],
        "publish_date": "", "publish_date_text": "2026年8月",
        "publish_date_precision": "month", "filename_date": "2026-08-14",
        "latest_date": "", "cutoff_scope": "mixed", "cutoff_pages": "7",
        "primary_topic": "海外宏观与大类资产展望",
        "topic_tags": "海外宏观|美债|美元|商品|权益|跨资产",
        "report_type": "cross_asset_macro", "research_horizon": "2026H2",
        "instrument_scope": "美债现券|美元|商品|权益|跨资产",
        "chart_density": "high", "extraction_quality": "medium", "overlap_group": "",
        "institution_page": 37,
        "notes": "[U] 封面仅明确2026年8月；p7仅写“截止8.7”且年份未明示；联合品牌关系待视觉确认。",
    },
    {
        "filename": "东证期货_半年报_国债_起伏相循，择时择势_张粲东_20260624.pdf",
        "sha256": "dadd32fd61400f1ac85220246812e4d2d53d29281b56516ff2290367a3142ad6",
        "page_count": 23, "title": "起伏相循，择时择势",
        "institution": "上海东证期货有限公司", "co_brand": "东证衍生品研究院",
        "authors": [("张粲东", "宏观策略高级分析师", 1)],
        "publish_date": "2026-06-23", "publish_date_text": "2026 年 6 月 23 日",
        "publish_date_precision": "day", "filename_date": "2026-06-24",
        "latest_date": "", "cutoff_scope": "mixed", "cutoff_pages": "",
        "primary_topic": "国债半年择时与曲线策略",
        "topic_tags": "宏观基本面|货币政策|收益率曲线|期限利差|期债",
        "report_type": "half_year_outlook", "research_horizon": "2026H2",
        "instrument_scope": "现券|收益率曲线|期限利差|国债期货",
        "chart_density": "high", "extraction_quality": "high", "overlap_group": "",
        "institution_page": 23,
        "notes": "[U] 封面日期2026-06-23与文件名日期2026-06-24冲突；PDF属性作者“王华柱”与署名冲突。",
    },
    {
        "filename": "光期宏观：2026年下半年国债报告.pdf",
        "sha256": "e194db58c4f9f223f50f5d122f0805e0d998ecae9a5ba2bdb49e6b5b96c8524c",
        "page_count": 30, "title": "2026年下半年国债报告",
        "institution": "光大期货有限公司", "co_brand": "光大期货研究所",
        "authors": [],
        "publish_date": "2026-06-28", "publish_date_text": "2026年06月28日",
        "publish_date_precision": "day", "filename_date": "",
        "latest_date": "2026-06-26", "cutoff_scope": "mixed", "cutoff_pages": "3,8",
        "primary_topic": "2026年下半年国债展望",
        "topic_tags": "宏观基本面|货币政策|收益率曲线|国债期货",
        "report_type": "half_year_outlook", "research_horizon": "2026H2",
        "instrument_scope": "现券|收益率曲线|国债期货",
        "chart_density": "high", "extraction_quality": "medium", "overlap_group": "",
        "institution_page": 30,
        "notes": "[U] 未发现明确报告作者；p29朱金涛简介不足以证明其为本篇作者。PDF属性为模板噪声。",
    },
]


PILOTS = [
    ("A", "综合半年报", "A_comprehensive_halfyear", "A", 1, "20260706-华泰期货-国债半年报：反弹之后，等待拐点.pdf", "华泰期货", "8b3b504abd2c6842"),
    ("C", "机构行为/基金", "C_institution_fund", "A", 2, "20260810-国泰海通证券-债基久期高位之后：关注绩优基金久期“死叉”.pdf", "国泰海通证券", "08caab59cd7e85db"),
    ("E", "期货微观结构/期现", "E_futures_micro_cash", "A", 3, "【期债半年报】水活则鱼动，把握流动性改善的机遇.pdf", "中信建投期货", "246bfde0ea4c6555"),
    ("B", "周度市场", "B_weekly_market", "B", 4, "20260809-华泰期货-国债周报：国债期货偏弱震荡，长端跌幅大于短端.pdf", "华泰期货", "d838cbb390787fce"),
    ("D", "事件点评", "D_event_commentary", "B", 5, "20260730-国泰海通证券-7月FOMC：沃什说了什么，票委做了什么.pdf", "国泰海通证券", "0177d692defb0aef"),
    ("F", "海外宏观/跨境", "F_overseas_crossborder", "B", 6, "20260814-国泰君安期货-2026H2海外宏观经济及大类资产展望：旧序余寒，新质去伪存真.pdf", "国泰君安期货", "67d92e32095e0725"),
]


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_csv(path: Path, columns: list[str], rows: list[dict[str, object]]) -> None:
    if path.exists():
        raise FileExistsError(f"Refusing to overwrite initialized registry: {path}")
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="raise", lineterminator="\r\n")
        writer.writeheader()
        writer.writerows(rows)


def page_bounds(page_spec: str) -> tuple[object, object]:
    if not page_spec:
        return "", ""
    cleaned = page_spec.replace(",", "-")
    parts = cleaned.split("-")
    return int(parts[0]), int(parts[-1])


def report_id(seed: dict[str, object]) -> str:
    return "rpt_" + str(seed["sha256"])[:16]


def evidence_row(counter: int, rid: str, field: str, value: object, klass: str,
                 source_type: str, page_start: object = "", page_end: object = "",
                 evidence_cutoff: str = "", cutoff_scope: str = "none",
                 method: str = "manual_text_review", confidence: str = "high",
                 note: str = "") -> dict[str, object]:
    return {
        "evidence_id": f"ev_20260815_{counter:04d}", "report_id": rid,
        "field_name": field, "field_value": value, "evidence_class": klass,
        "source_type": source_type, "page_start": page_start, "page_end": page_end,
        "evidence_data_cutoff": evidence_cutoff, "cutoff_scope": cutoff_scope,
        "extraction_method": method, "confidence": confidence,
        "review_status": "pending", "note": note,
    }


def main() -> None:
    required_targets = [ROOT / "01_registry" / "reports.csv", ROOT / "01_registry" / "report_metadata_evidence.csv"]
    if any(path.exists() for path in required_targets):
        raise SystemExit("Registry already exists; initialization will not overwrite it.")
    script_hash = sha256(Path(__file__))
    report_rows: list[dict[str, object]] = []
    evidence_rows: list[dict[str, object]] = []
    issue_rows: list[dict[str, object]] = []
    processing_rows: list[dict[str, object]] = []
    researcher_sources: dict[tuple[str, str], dict[str, object]] = {}
    evidence_counter = 0
    issue_counter = 0
    step_no = 0
    report_ids: set[str] = set()

    for seed in REPORTS:
        source = SOURCE_ROOT / str(seed["filename"])
        target = RAW_REPORTS / str(seed["filename"])
        if not source.exists() or not target.exists():
            raise FileNotFoundError(f"Missing source or target: {seed['filename']}")
        source_hash, target_hash = sha256(source), sha256(target)
        if source_hash != seed["sha256"] or target_hash != source_hash:
            raise ValueError(f"SHA mismatch: {seed['filename']}")
        reader = PdfReader(str(target))
        if len(reader.pages) != seed["page_count"]:
            raise ValueError(f"Page-count mismatch: {seed['filename']}")
        texts: list[str] = []
        for page in reader.pages:
            try:
                texts.append(page.extract_text() or "")
            except Exception:
                texts.append("")
        text_pages = sum(bool(text.strip()) for text in texts)
        rid = report_id(seed)
        if rid in report_ids:
            raise ValueError(f"report_id prefix collision: {rid}")
        report_ids.add(rid)
        authors_display = "|".join(
            name if role == "作者姓名" else f"{name}（{role}）"
            for name, role, _ in seed["authors"]
        )
        duplicate_status = "partial_overlap_review" if seed["overlap_group"] else "unique"
        report_rows.append({
            "report_id": rid,
            "relative_path": target.relative_to(ROOT).as_posix(),
            "source_original_path": str(source),
            "filename": seed["filename"], "sha256": source_hash,
            "file_size_bytes": target.stat().st_size, "page_count": len(reader.pages),
            "title": seed["title"], "institution": seed["institution"],
            "co_brand": seed["co_brand"], "authors_display": authors_display,
            "publish_date": seed["publish_date"], "publish_date_text": seed["publish_date_text"],
            "publish_date_precision": seed["publish_date_precision"],
            "filename_date": seed["filename_date"], "data_cutoff": "",
            "latest_explicit_data_date": seed["latest_date"],
            "data_cutoff_scope": seed["cutoff_scope"], "primary_topic": seed["primary_topic"],
            "topic_tags": seed["topic_tags"], "report_type": seed["report_type"],
            "research_horizon": seed["research_horizon"], "instrument_scope": seed["instrument_scope"],
            "has_text_layer": "true", "text_page_coverage_pct": round(100 * text_pages / len(reader.pages), 2),
            "text_char_count": sum(len(text) for text in texts),
            "ocr_requirement": "none" if seed["chart_density"] == "medium" else "targeted",
            "visual_review_requirement": "required", "chart_density": seed["chart_density"],
            "duplicate_status": duplicate_status, "duplicate_of_report_id": "",
            "partial_overlap_group": seed["overlap_group"],
            "extraction_quality": seed["extraction_quality"], "manual_review_status": "pending",
            "ingested_at": INGESTED_AT, "notes": seed["notes"],
        })
        identity_note = "身份元数据不适用数据截止日；以U状态说明，不以发布日期代替。"
        for field, value, klass, source_type, p1, p2, method, note in [
            ("sha256", source_hash, "X", "filesystem", "", "", "sha256", "源与项目副本SHA-256一致。"),
            ("page_count", len(reader.pages), "X", "pdf_structure", 1, len(reader.pages), "pypdf", "PDF物理页数。"),
            ("has_text_layer", "true", "X", "pdf_text_layer", 1, len(reader.pages), "pypdf_text_extract", f"{text_pages}/{len(reader.pages)}页存在可提取文字。"),
            ("title", seed["title"], "E", "pdf_text", 1, 1, "manual_text_review", identity_note),
            ("institution", seed["institution"], "E", "pdf_text", seed["institution_page"], seed["institution_page"], "manual_text_review", identity_note),
            ("publish_date_text", seed["publish_date_text"], "E", "pdf_text", 1, 1, "manual_text_review", identity_note),
        ]:
            evidence_counter += 1
            evidence_rows.append(evidence_row(evidence_counter, rid, field, value, klass, source_type, p1, p2, method=method, note=note))
        if seed["publish_date"]:
            evidence_counter += 1
            evidence_rows.append(evidence_row(evidence_counter, rid, "publish_date", seed["publish_date"], "E", "pdf_text", 1, 1, method="date_normalization", note=identity_note))
        else:
            evidence_counter += 1
            evidence_rows.append(evidence_row(evidence_counter, rid, "publish_date", "", "U", "pdf_text", 1, 1, method="manual_text_review", confidence="medium", note="封面只明确月份；精确日期不得由文件名覆盖。"))
        if seed["filename_date"]:
            evidence_counter += 1
            evidence_rows.append(evidence_row(evidence_counter, rid, "filename_date", seed["filename_date"], "X", "filename", method="filename_parse", note="仅为文件名日期，不等同报告发布日期或数据截止日。"))
        if seed["co_brand"]:
            evidence_counter += 1
            evidence_rows.append(evidence_row(evidence_counter, rid, "co_brand", seed["co_brand"], "E", "pdf_text", 1, 1, method="manual_text_review", note=identity_note))
        elif rid == "rpt_bb2de5c3b82fc0e2":
            evidence_counter += 1
            evidence_rows.append(evidence_row(evidence_counter, rid, "co_brand", "", "U", "pdf_visual_review_required", 1, 1, method="text_layer_review", confidence="low", note="联合品牌关系未能在文字层可靠确认，保留空值。"))
        if seed["authors"]:
            author_start = min(page for _, _, page in seed["authors"])
            author_end = max(page for _, _, page in seed["authors"])
            evidence_counter += 1
            evidence_rows.append(evidence_row(evidence_counter, rid, "authors_display", authors_display, "E", "pdf_text", author_start, author_end, method="manual_text_review", note=identity_note))
            for name, role, page in seed["authors"]:
                key = (name, str(seed["institution"]))
                entry = researcher_sources.setdefault(key, {"name": name, "institution": seed["institution"], "roles": set(), "reports": [], "pages": []})
                entry["roles"].add(role)
                entry["reports"].append(rid)
                entry["pages"].append(f"{rid}:p{page}")
        else:
            evidence_counter += 1
            evidence_rows.append(evidence_row(evidence_counter, rid, "authors_display", "", "U", "pdf_text", 29, 29, method="manual_text_review", confidence="low", note="朱金涛仅出现在研究员简介页，不足以证明其为本篇作者。"))
        if seed["latest_date"]:
            p1, p2 = page_bounds(str(seed["cutoff_pages"]))
            evidence_counter += 1
            evidence_rows.append(evidence_row(evidence_counter, rid, "latest_explicit_data_date", seed["latest_date"], "E", "pdf_text_or_chart_axis", p1, p2, str(seed["latest_date"]), str(seed["cutoff_scope"]), "manual_text_and_layout_review", "medium", "仅代表报告内最新可明确核证的局部日期。"))
        elif rid == "rpt_bb2de5c3b82fc0e2":
            evidence_counter += 1
            evidence_rows.append(evidence_row(evidence_counter, rid, "latest_explicit_data_date", "截止8.7（年份未明示）", "U", "pdf_text", 7, 7, cutoff_scope="mixed", method="manual_text_review", confidence="low", note="不得推断为2026-08-07写入ISO字段。"))
        else:
            evidence_counter += 1
            evidence_rows.append(evidence_row(evidence_counter, rid, "latest_explicit_data_date", "", "U", "pdf_text", cutoff_scope=str(seed["cutoff_scope"]), method="manual_text_review", confidence="medium", note="未找到可安全标准化的最新全局日期。"))
        evidence_counter += 1
        evidence_rows.append(evidence_row(evidence_counter, rid, "data_cutoff", "", "U", "pdf_text_and_layout_review", cutoff_scope=str(seed["cutoff_scope"]), method="manual_review", confidence="high", note="未发现全篇统一数据截止日；不得用发布日期或最新局部日期补齐。"))
        for field in ("primary_topic", "report_type", "chart_density"):
            evidence_counter += 1
            evidence_rows.append(evidence_row(evidence_counter, rid, field, seed[field], "I", "full_report_review", 1, len(reader.pages), cutoff_scope=str(seed["cutoff_scope"]), method="controlled_vocabulary_mapping", confidence="medium", note="研究资料管理员的分类映射，非报告明示事实。"))

        issue_counter += 1
        issue_rows.append({
            "issue_id": f"iss_20260815_{issue_counter:04d}", "report_id": rid, "asset_id": "",
            "field_name": "chart_values", "page_start": 1, "page_end": len(reader.pages),
            "issue_type": "visual_chart_review_required", "severity": "medium", "evidence_class": "U",
            "description": "文字层存在不代表曲线点位、表格行列、图例方向、历史分位和单位可可靠恢复；写入[E]前需逐图视觉复核。",
            "status": "open", "resolution": "", "detected_at": INGESTED_AT, "reviewed_at": "",
        })
        if seed["cutoff_scope"] != "report_wide":
            issue_counter += 1
            issue_rows.append({
                "issue_id": f"iss_20260815_{issue_counter:04d}", "report_id": rid, "asset_id": "",
                "field_name": "data_cutoff", "page_start": "", "page_end": "",
                "issue_type": "report_wide_cutoff_missing_or_mixed", "severity": "medium", "evidence_class": "U",
                "description": f"未发现全篇统一数据截止日；当前scope={seed['cutoff_scope']}，data_cutoff保持为空。",
                "status": "open", "resolution": "", "detected_at": INGESTED_AT, "reviewed_at": "",
            })
        if seed["overlap_group"]:
            issue_counter += 1
            issue_rows.append({
                "issue_id": f"iss_20260815_{issue_counter:04d}", "report_id": rid, "asset_id": "",
                "field_name": "duplicate_status", "page_start": 1, "page_end": len(reader.pages),
                "issue_type": "partial_overlap_review", "severity": "low", "evidence_class": "I",
                "description": f"登记于{seed['overlap_group']}，仅表示局部内容/模板复用待核，不构成删除依据。",
                "status": "open", "resolution": "", "detected_at": INGESTED_AT, "reviewed_at": "",
            })
        tailored = {
            "rpt_c406ce66269f6c16": ("pdf_metadata", "metadata_conflict", "high", "PDF属性标题为模板名、作者为“1”，与封面/正文冲突。"),
            "rpt_bb2de5c3b82fc0e2": ("publish_date|latest_explicit_data_date|co_brand", "precision_and_brand_uncertain", "high", "精确发布日期、‘截止8.7’年份和联合品牌关系需要视觉/人工确认。"),
            "rpt_dadd32fd61400f1a": ("publish_date|filename_date|authors_display", "metadata_conflict", "high", "封面日期与文件名日期冲突；PDF属性作者与明确署名冲突。"),
            "rpt_e194db58c4f9f223": ("authors_display|pdf_metadata", "author_and_metadata_uncertain", "high", "未明确署名；研究员简介不得自动当作者，PDF属性为模板噪声。"),
        }.get(rid)
        if tailored:
            field_name, issue_type, severity, description = tailored
            issue_counter += 1
            issue_rows.append({
                "issue_id": f"iss_20260815_{issue_counter:04d}", "report_id": rid, "asset_id": "",
                "field_name": field_name, "page_start": "", "page_end": "",
                "issue_type": issue_type, "severity": severity, "evidence_class": "U",
                "description": description, "status": "open", "resolution": "",
                "detected_at": INGESTED_AT, "reviewed_at": "",
            })
        step_no += 1
        processing_rows.append({
            "run_id": RUN_ID, "step_no": step_no, "timestamp": INGESTED_AT,
            "object_type": "report_pdf", "object_id": rid, "stage": "ingest",
            "action": "copy_verify_readonly", "tool": "Copy-Item+pypdf+hashlib",
            "tool_version": f"python {sys.version_info.major}.{sys.version_info.minor}",
            "input_sha256": source_hash, "output_relative_path": target.relative_to(ROOT).as_posix(),
            "output_sha256": target_hash, "status": "success", "records_or_pages": len(reader.pages),
            "config_sha256": script_hash, "git_commit": "",
            "message": "源/目标SHA与页数一致；项目副本设置Windows只读。",
        })

    report_rows.sort(key=lambda row: str(row["filename"]))
    evidence_rows.sort(key=lambda row: str(row["evidence_id"]))
    issue_rows.sort(key=lambda row: str(row["issue_id"]))

    workbook = PENDING_REVIEW / "20260811-国泰海通证券-基金短端增持降温——信用债机构久期热度观测.xlsx"
    source_workbook = SOURCE_ROOT / workbook.name
    workbook_hash = sha256(workbook)
    if workbook_hash != sha256(source_workbook) or workbook_hash != "ef8d6cff391f5ab0beebe5b77c6b48380ea261d70625891456370c10473901fd":
        raise ValueError("Workbook SHA mismatch")
    wb = load_workbook(workbook, read_only=True, data_only=False)
    wb_values = load_workbook(workbook, read_only=True, data_only=True)
    asset_id = "asset_" + workbook_hash[:16]
    asset_rows = [{
        "asset_id": asset_id, "relative_path": workbook.relative_to(ROOT).as_posix(),
        "source_original_path": str(source_workbook), "filename": workbook.name,
        "file_type": "xlsx", "sha256": workbook_hash, "file_size_bytes": workbook.stat().st_size,
        "asset_role": "research_workbook", "associated_report_id": "",
        "institution": "国泰海通证券股份有限公司", "asset_date": "2026-08-11",
        "sheet_count": len(wb.sheetnames), "extraction_quality": "medium",
        "manual_review_status": "pending",
        "notes": "[X] 机构和日期来自文件名，待人工确认；首页!B59为#VALUE!错误常量，OOXML无可恢复公式节点。",
    }]
    issue_counter += 1
    issue_rows.append({
        "issue_id": f"iss_20260815_{issue_counter:04d}", "report_id": "", "asset_id": asset_id,
        "field_name": "首页!B59", "page_start": "", "page_end": "",
        "issue_type": "spreadsheet_error_value_no_formula", "severity": "high", "evidence_class": "U",
        "description": f"单元格存储值为{wb_values['首页']['B59'].value}，当前OOXML没有可恢复公式节点；需回到上游模板或生成链路人工复核。",
        "status": "open", "resolution": "", "detected_at": INGESTED_AT, "reviewed_at": "",
    })
    step_no += 1
    processing_rows.append({
        "run_id": RUN_ID, "step_no": step_no, "timestamp": INGESTED_AT,
        "object_type": "source_asset", "object_id": asset_id, "stage": "ingest",
        "action": "copy_verify_pending_review", "tool": "Copy-Item+openpyxl+hashlib",
        "tool_version": f"python {sys.version_info.major}.{sys.version_info.minor}",
        "input_sha256": workbook_hash, "output_relative_path": workbook.relative_to(ROOT).as_posix(),
        "output_sha256": workbook_hash, "status": "success", "records_or_pages": len(wb.sheetnames),
        "config_sha256": script_hash, "git_commit": "",
        "message": "工作簿作为独立附件登记；未混入PDF报告目录。",
    })

    researcher_rows: list[dict[str, object]] = []
    for (name, institution), entry in sorted(researcher_sources.items()):
        researcher_rows.append({
            "researcher_id": "rsr_" + hashlib.sha256(f"{name}|{institution}".encode("utf-8")).hexdigest()[:16],
            "name": name, "institution": institution,
            "roles_display": "|".join(sorted(entry["roles"])),
            "source_report_ids": "|".join(dict.fromkeys(entry["reports"])),
            "source_pages": "|".join(dict.fromkeys(entry["pages"])),
            "evidence_class": "E", "manual_review_status": "pending",
            "notes": "仅录入封面或正文明确署名；未明确的作者角色不扩写。",
        })

    rid_by_filename = {row["filename"]: row["report_id"] for row in report_rows}
    pilot_rows: list[dict[str, object]] = []
    for stratum, name, hash_stratum_key, gate, order, filename, family, expected_prefix in PILOTS:
        key = f"20260815|{hash_stratum_key}|{unicodedata.normalize('NFC', filename)}"
        selection_hash = hashlib.sha256(key.encode("utf-8")).hexdigest()
        if not selection_hash.startswith(expected_prefix):
            raise ValueError(f"Pilot hash mismatch for {filename}: {selection_hash}")
        pilot_rows.append({
            "selection_date": "2026-08-15", "seed": 20260815,
            "stratum_code": stratum, "stratum_name": name,
            "hash_stratum_key": hash_stratum_key, "gate": f"Gate {gate}",
            "gate_order": order, "filename": filename, "report_id": rid_by_filename[filename],
            "publisher_template_family": family, "selection_hash": selection_hash,
            "selection_hash_prefix": expected_prefix, "family_cap": 2,
            "selection_status": "planned_not_processed",
            "user_approval_required": "true",
            "notes": "Gate A经用户明确审阅通过前不得生成Gate B卡片；当前不生成任何摘要或卡片。",
        })

    registry_dir = ROOT / "01_registry"
    audit_dir = ROOT / "09_audit"
    write_csv(registry_dir / "reports.csv", REPORT_COLUMNS, report_rows)
    write_csv(registry_dir / "report_metadata_evidence.csv", EVIDENCE_COLUMNS, evidence_rows)
    write_csv(registry_dir / "source_assets.csv", ASSET_COLUMNS, asset_rows)
    write_csv(registry_dir / "researchers.csv", [
        "researcher_id", "name", "institution", "roles_display", "source_report_ids",
        "source_pages", "evidence_class", "manual_review_status", "notes",
    ], researcher_rows)
    write_csv(audit_dir / "extraction_issues.csv", ISSUE_COLUMNS, issue_rows)
    write_csv(audit_dir / "pilot_selection_20260815.csv", [
        "selection_date", "seed", "stratum_code", "stratum_name", "hash_stratum_key", "gate", "gate_order",
        "filename", "report_id", "publisher_template_family", "selection_hash",
        "selection_hash_prefix", "family_cap", "selection_status",
        "user_approval_required", "notes",
    ], pilot_rows)

    empty_csvs = [
        (ROOT / "04_indicators" / "indicator_dictionary.csv", [
            "indicator_id", "indicator_name", "category", "definition", "unit", "frequency",
            "source_id", "source_field", "calculation_rule", "availability_lag",
            "market_object", "evidence_class", "status", "notes",
        ]),
        (ROOT / "04_indicators" / "data_source_map.csv", [
            "source_id", "source_name", "provider", "interface", "endpoint_or_template",
            "credential_target", "source_priority", "update_frequency", "raw_path_pattern",
            "processed_path_pattern", "license_or_usage_note", "status", "notes",
        ]),
        (ROOT / "06_weekly" / "view_ledger.csv", [
            "view_id", "view_version", "created_at", "weekly_id", "statement",
            "evidence_class", "evidence_ids", "applicable_horizon", "transmission_mechanism",
            "trigger_conditions", "falsification_conditions", "tracking_indicators",
            "market_pricing", "current_status", "prior_view_id", "review_result",
            "reviewed_at", "owner_decision", "notes",
        ]),
        (ROOT / "07_events" / "policy_events.csv", [
            "event_id", "event_date", "event_time", "timezone", "event_type", "authority",
            "title", "fact", "evidence_class", "source", "source_date",
            "affected_market_objects", "transmission_mechanism", "trigger_conditions",
            "falsification_conditions", "tracking_indicators", "review_status", "notes",
        ]),
        (ROOT / "07_events" / "economic_calendar.csv", [
            "event_id", "scheduled_date", "scheduled_time", "timezone", "country_region",
            "indicator", "period", "prior", "consensus", "actual", "unit", "source",
            "source_date", "evidence_class", "affected_market_objects", "review_status", "notes",
        ]),
        (ROOT / "07_events" / "market_anomalies.csv", [
            "anomaly_id", "event_date", "market_object", "metric", "observed_value", "unit",
            "comparison_baseline", "threshold_rule", "fact", "interpretation",
            "evidence_class", "source", "source_date", "trigger_conditions",
            "falsification_conditions", "tracking_indicators", "review_status", "notes",
        ]),
        (audit_dir / "forecast_scorecard.csv", [
            "forecast_id", "view_id", "forecast_date", "applicable_horizon", "target_object",
            "forecast_statement", "trigger_conditions", "falsification_conditions",
            "tracking_indicators", "evaluation_date", "realized_outcome", "score_method",
            "score", "review_status", "notes",
        ]),
        (audit_dir / "data_migration_issues.csv", [
            "issue_id", "stage", "source_path", "target_relative_path", "issue_type",
            "severity", "evidence_class", "description", "status", "blocking_condition",
            "resolution", "detected_at", "reviewed_at",
        ]),
    ]
    for path, columns in empty_csvs:
        write_csv(path, columns, [])

    for path in [
        registry_dir / "reports.csv", registry_dir / "report_metadata_evidence.csv",
        registry_dir / "source_assets.csv", audit_dir / "extraction_issues.csv",
        audit_dir / "pilot_selection_20260815.csv",
    ]:
        step_no += 1
        processing_rows.append({
            "run_id": RUN_ID, "step_no": step_no, "timestamp": INGESTED_AT,
            "object_type": "registry", "object_id": path.stem, "stage": "registry_initialize",
            "action": "create_new", "tool": "initialize_registry.py",
            "tool_version": f"python {sys.version_info.major}.{sys.version_info.minor}",
            "input_sha256": "", "output_relative_path": path.relative_to(ROOT).as_posix(),
            "output_sha256": sha256(path), "status": "success",
            "records_or_pages": sum(1 for _ in path.open("r", encoding="utf-8-sig", newline="")) - 1,
            "config_sha256": script_hash, "git_commit": "",
            "message": "初始化新文件；未生成报告摘要或报告卡片。",
        })
    write_csv(registry_dir / "processing_log.csv", PROCESSING_COLUMNS, processing_rows)

    print(json.dumps({
        "reports": len(report_rows), "evidence_rows": len(evidence_rows),
        "researchers": len(researcher_rows), "issues": len(issue_rows),
        "source_assets": len(asset_rows), "pilots": len(pilot_rows),
        "ingested_at": INGESTED_AT,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
