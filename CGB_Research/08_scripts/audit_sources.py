"""Read-only source inventory used for initialization and later integrity checks."""

from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def pdf_record(path: Path) -> dict[str, object]:
    reader = PdfReader(str(path))
    texts: list[str] = []
    for page in reader.pages:
        try:
            texts.append(page.extract_text() or "")
        except Exception:
            texts.append("")
    stat = path.stat()
    return {
        "filename": path.name,
        "sha256": sha256(path),
        "size": stat.st_size,
        "pages": len(reader.pages),
        "encrypted": bool(reader.is_encrypted),
        "text_pages": sum(bool(text.strip()) for text in texts),
        "text_chars": sum(len(text) for text in texts),
        "mtime_ns": stat.st_mtime_ns,
    }


def workbook_record(path: Path) -> dict[str, object]:
    formulas = load_workbook(path, read_only=True, data_only=False)
    values = load_workbook(path, read_only=True, data_only=True)
    stat = path.stat()
    return {
        "filename": path.name,
        "sha256": sha256(path),
        "size": stat.st_size,
        "sheet_count": len(formulas.sheetnames),
        "sheets": formulas.sheetnames,
        "首页_B59_formula": formulas["首页"]["B59"].value,
        "首页_B59_cached": values["首页"]["B59"].value,
        "mtime_ns": stat.st_mtime_ns,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source_dir", type=Path)
    args = parser.parse_args()
    pdfs = sorted(args.source_dir.glob("*.pdf"), key=lambda path: path.name)
    workbooks = sorted(args.source_dir.glob("*.xlsx"), key=lambda path: path.name)
    payload = {
        "pdfs": [pdf_record(path) for path in pdfs],
        "workbooks": [workbook_record(path) for path in workbooks],
    }
    print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
